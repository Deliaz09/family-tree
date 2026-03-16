import io
import json
import os
import sys
import uuid
import unicodedata
import shutil
import secrets
import traceback
import tempfile
import threading
from datetime import datetime, timezone, timedelta
from difflib import SequenceMatcher
from concurrent.futures import ThreadPoolExecutor
from typing import List
from urllib.parse import urlparse

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Depends, Header, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from neo4j import GraphDatabase

from gedcom_utils import parseaza_gedcom, genereaza_gedcom
from email_utils import trimite_email
from rudenie_utils import decodeaza_drum, numara_alianta, construieste_raspuns
from validari_utils import (
    eroare_deces_inainte_nastere, avertismente_varste_parinte,
    este_descendent, linie_directa, gaseste_cicluri,
    normalizeaza_nume,
)

load_dotenv()
load_dotenv(os.path.join("..", ".env"))

NEO4J_URI       = os.getenv("NEO4J_URI")
NEO4J_USER      = os.getenv("NEO4J_USER", "neo4j")
NEO4J_PASS      = os.getenv("NEO4J_PASSWORD")
SUPABASE_URL    = os.getenv("SUPABASE_URL", "")
SUPABASE_ANON   = os.getenv("SUPABASE_ANON_KEY", "")
SUPABASE_SERVICE = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
SUPABASE_BUCKET = os.getenv("SUPABASE_PHOTOS_BUCKET", "photos")
APP_PUBLIC_URL  = os.getenv("APP_PUBLIC_URL", "")
FRONTEND_URL    = os.getenv("FRONTEND_URL", "http://localhost:5173")
FRONTEND_PORT   = os.getenv("FRONTEND_PORT", "5173")

def slug_nume_fisier(nume: str) -> str:
    if not nume:
        return nume
    nfkd = unicodedata.normalize("NFKD", nume)
    fara_semne = "".join(c for c in nfkd if not unicodedata.combining(c))
    return "".join(c if ord(c) < 128 else "_" for c in fara_semne)

def slug_photo_url(u: str) -> str:
    if not u or all(ord(c) < 128 for c in u):
        return u
    idx = u.rfind("/")
    if idx == -1:
        return slug_nume_fisier(u)
    return u[:idx + 1] + slug_nume_fisier(u[idx + 1:])

def frontend_url(request: Request | None = None) -> str:
    configured = (APP_PUBLIC_URL or FRONTEND_URL or "").strip().rstrip("/")
    if configured:
        try:
            parsed = urlparse(configured)
            if parsed.hostname not in {"localhost", "127.0.0.1", "::1"}:
                return configured
        except Exception:
            return configured

    if request is not None:
        host = request.url.hostname or "localhost"
        scheme = request.url.scheme or "http"
        return f"{scheme}://{host}:{FRONTEND_PORT}"

    return configured or f"http://localhost:{FRONTEND_PORT}"

if not NEO4J_URI or not NEO4J_PASS:
    print("Credentialele Neo4j lipsesc din .env")
    sys.exit(1)

def doar_autentificat(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        print(f"[AUTH DEBUG] lipseste antetul Authorization (valoare={authorization!r})")
        raise HTTPException(401, "Token lipsa")

    token = authorization.split(" ")[1]

    if not SUPABASE_URL or not SUPABASE_ANON:
        raise HTTPException(500, "Configurarea Supabase lipseste pe server")

    try:
        import httpx
        with httpx.Client(timeout=10) as client:
            res = client.get(
                f"{SUPABASE_URL}/auth/v1/user",
                headers={
                    "apikey":        SUPABASE_ANON,
                    "Authorization": f"Bearer {token}",
                }
            )
        if res.status_code != 200:
            print(f"[AUTH DEBUG] Supabase a respins tokenul: status={res.status_code} "
                  f"len_token={len(token)} body={res.text[:200]!r}")
            raise HTTPException(401, "Token invalid sau expirat")
        u = res.json()
        meta = u.get("user_metadata") or {}
        return {
            "user_id":   u["id"],
            "email":     u.get("email", ""),
            "full_name": meta.get("full_name") or u.get("email", ""),
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[AUTH ERROR] {e}")
        raise HTTPException(401, "Eroare validare token")

def _supabase_admin_headers() -> dict:
    return {
        "apikey":        SUPABASE_SERVICE,
        "Authorization": f"Bearer {SUPABASE_SERVICE}",
        "Content-Type":  "application/json",
    }

def supabase_admin_creeaza_user(email: str, password: str, full_name: str) -> str:
    if not SUPABASE_URL or not SUPABASE_SERVICE:
        raise HTTPException(503, "Înregistrarea prin email nu este configurată pe server "
                                 "(lipsește SUPABASE_SERVICE_ROLE_KEY).")
    import httpx
    with httpx.Client(timeout=15) as client:
        res = client.post(
            f"{SUPABASE_URL}/auth/v1/admin/users",
            headers=_supabase_admin_headers(),
            json={
                "email": email,
                "password": password,
                "email_confirm": False,
                "user_metadata": {"full_name": full_name or email},
            },
        )
    if res.status_code in (200, 201):
        return res.json().get("id")
    txt = res.text.lower()
    if res.status_code in (422, 409) or "already" in txt or "exists" in txt or "registered" in txt:
        raise HTTPException(409, "Există deja un cont cu acest email.")
    print(f"[AUTH ADMIN] creare user esuata {res.status_code}: {res.text}")
    raise HTTPException(502, "Contul nu a putut fi creat.")

def supabase_admin_confirma_user(user_id: str) -> bool:
    if not SUPABASE_URL or not SUPABASE_SERVICE:
        return False
    import httpx
    with httpx.Client(timeout=15) as client:
        res = client.put(
            f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}",
            headers=_supabase_admin_headers(),
            json={"email_confirm": True},
        )
    if res.status_code == 200:
        return True
    print(f"[AUTH ADMIN] confirmare esuata {res.status_code}: {res.text}")
    return False

def creeaza_token_email(user_id: str, email: str, full_name: str, kind: str = "verify") -> str:
    token = secrets.token_urlsafe(32)
    acum  = datetime.now(timezone.utc)
    with driver.session() as s:
        s.run(
            "MATCH (t:EmailToken {email:$em, kind:$kind, used:false}) SET t.used=true",
            em=email.lower(), kind=kind,
        )
        s.run(
            "CREATE (t:EmailToken {token:$tok, user_id:$uid, email:$em, full_name:$nm, "
            "kind:$kind, used:false, created_at:$now, expires_at:$exp})",
            tok=token, uid=user_id, em=email.lower(), nm=full_name or email, kind=kind,
            now=acum.isoformat(), exp=(acum + timedelta(hours=24)).isoformat(),
        )
    return token

def supabase_admin_seteaza_parola(user_id: str, password: str) -> bool:
    if not SUPABASE_URL or not SUPABASE_SERVICE:
        return False
    import httpx
    with httpx.Client(timeout=15) as client:
        res = client.put(
            f"{SUPABASE_URL}/auth/v1/admin/users/{user_id}",
            headers=_supabase_admin_headers(),
            json={"password": password},
        )
    if res.status_code == 200:
        return True
    print(f"[AUTH ADMIN] setare parola esuata {res.status_code}: {res.text}")
    return False

def gaseste_user_id_dupa_email(email: str):
    em = email.lower()
    with driver.session() as s:
        row = s.run(
            "MATCH (u:AppUser) WHERE toLower(u.email)=$em RETURN u.user_id AS uid LIMIT 1",
            em=em,
        ).single()
        if row and row["uid"]:
            return row["uid"]
        row = s.run(
            "MATCH (t:EmailToken {email:$em}) RETURN t.user_id AS uid "
            "ORDER BY t.created_at DESC LIMIT 1",
            em=em,
        ).single()
        return row["uid"] if row else None

def gaseste_user_id_supabase_dupa_email(email: str):
    if not SUPABASE_URL or not SUPABASE_SERVICE:
        return None

    import httpx

    headers = _supabase_admin_headers()
    headers.pop("Content-Type", None)

    with httpx.Client(timeout=15) as client:
        page = 1
        while True:
            res = client.get(
                f"{SUPABASE_URL}/auth/v1/admin/users",
                headers=headers,
                params={"page": page, "per_page": 100},
            )
            if res.status_code != 200:
                print(f"[AUTH ADMIN] listare users esuata {res.status_code}: {res.text}")
                return None

            data = res.json()
            users = data.get("users") or []
            if not users:
                return None

            for user in users:
                if (user.get("email") or "").lower() == em:
                    return user.get("id")

            total_pages = data.get("last_page")
            if total_pages and page >= int(total_pages):
                return None
            if len(users) < 100:
                return None
            page += 1

def trimite_email_verificare(email: str, full_name: str, token: str, request: Request | None = None) -> bool:
    link = f"{frontend_url(request)}/verify-email?token={token}"
    return trimite_email(
        to_email=email,
        to_name=full_name or email,
        subject="Confirmă-ți adresa de email — Arbore Genealogic",
        heading="Bine ai venit!",
        icon="🌳",
        message=(
            "Îți mulțumim că ți-ai creat cont la Arbore Genealogic — locul unde "
            "construiești și păstrezi povestea familiei tale.\n\n"
            "Mai e un singur pas: confirmă-ți adresa de email apăsând pe butonul de mai jos. "
            "Linkul este valabil 24 de ore.\n\n"
            "Dacă nu tu ai creat acest cont, poți ignora în siguranță acest mesaj."
        ),
        link=link,
        link_label="Confirmă adresa de email",
    )

def _rol_membru(tree_id: str, caller_id: str, email: str):
    em = (email or "").lower()
    with driver.session() as s:
        row = s.run(
            '\n            MATCH (m:Member {tree_id:$tid})\n            WHERE m.user_id = $uid\n               OR (m.email IS NOT NULL AND toLower(m.email) = $em)\n            RETURN m.role AS role, m.user_id AS muid, m.email AS memail\n            LIMIT 1\n            ',
            tid=tree_id, uid=caller_id, em=em,
        ).single()
        if not row:
            return None

        role = row["role"] or "editor"
        if role == "admin":
            role = "editor"
            s.run(
                "MATCH (m:Member {tree_id:$tid}) WHERE m.user_id=$uid OR toLower(m.email)=$em "
                "SET m.role='editor'",
                tid=tree_id, uid=caller_id, em=em,
            )

        if not row["muid"] and row["memail"]:
            s.run(
                "MATCH (m:Member {tree_id:$tid}) WHERE toLower(m.email)=$em "
                "SET m.user_id=$uid, m.status='active'",
                tid=tree_id, em=row["memail"].lower(), uid=caller_id,
            )
        return role

def utilizator_curent(
    cu0: dict = Depends(doar_autentificat),
    x_tree_id: str = Header(None),
) -> dict:
    caller   = cu0["user_id"]
    email    = cu0["email"]
    nume     = cu0["full_name"]
    tree_id  = x_tree_id or caller

    if tree_id == caller:
        role = "owner"
    else:
        role = _rol_membru(tree_id, caller, email)
        if role is None:
            raise HTTPException(403, "Nu ai acces la acest arbore.")

    return {
        "user_id":   tree_id,
        "caller_id": caller,
        "role":      role,
        "email":     email,
        "full_name": nume,
    }

def cere_scriere(cu: dict):
    if cu.get("role") not in ("owner", "editor"):
        raise HTTPException(403, "Nu ai drept de editare pe acest arbore.")

def cere_owner(cu: dict):
    if cu.get("role") != "owner":
        raise HTTPException(403, "Doar proprietarul arborelui poate face aceasta operatie.")

def nou_id(uid: str, prefix: str = "p") -> str:
    return f"{uid}_{prefix}_{uuid.uuid4().hex[:8]}"

def genereaza_id_stabil(uid: str, original_id=None, full_name: str = "", birth=None) -> str:
    if original_id and str(original_id).strip():
        cheie = str(original_id).strip()
    else:
        nume_norm = " ".join(str(full_name).lower().split()) if full_name else "necunoscut"
        an = str(birth).strip() if birth else "0000"
        cheie = f"{nume_norm}_{an}"
    return f"{uid}_p_{cheie}"

def genereaza_id_relatie_stabil(uid: str, rel_id_original=None,
                                 male_pid=None, female_pid=None) -> str:
    if rel_id_original and str(rel_id_original).strip():
        cheie = str(rel_id_original).strip()
    else:
        m = male_pid or "none"
        f = female_pid or "none"
        cheie = f"{m}__{f}"
    return f"{uid}_r_{cheie}"

def serializeaza_persoana(node) -> dict:
    d = dict(node)
    if not d.get("photo_url") and d.get("photo"):
        d["photo_url"] = f"/photos/{d['photo']}"
    if d.get("photo_url"):
        d["photo_url"] = slug_photo_url(d["photo_url"])
    return d

def int_sau_none(val):
    try:
        return int(float(str(val).strip())) if val and str(val).strip() else None
    except (ValueError, TypeError):
        return None

def acum_iso():
    return datetime.now(timezone.utc).isoformat()

def json_dump(data):
    return json.dumps(data if data is not None else {}, ensure_ascii=False, default=str)

def audit_log(s, cu: dict, actiune: str, entitate: str, entitate_id: str = "",
              inainte=None, dupa=None, detalii=None):
    s.run(
        "CREATE (a:AuditLog {id:$id, tree_id:$tid, actor_id:$actor, actor_email:$email, "
        "actor_name:$name, role:$role, action:$action, entity_type:$etype, entity_id:$eid, "
        "before_json:$before, after_json:$after, details_json:$details, created_at:$created})",
        id=nou_id(cu["user_id"], "audit"), tid=cu["user_id"],
        actor=cu.get("caller_id") or cu.get("user_id"), email=cu.get("email") or "",
        name=cu.get("full_name") or "", role=cu.get("role") or "",
        action=actiune, etype=entitate, eid=entitate_id or "",
        before=json_dump(inainte), after=json_dump(dupa), details=json_dump(detalii),
        created=acum_iso(),
    )

def mini_persoana(p):
    if not p:
        return None
    keys = ("id", "full_name", "given_name", "surname", "gender", "birth", "death",
            "photo_url", "address", "tel", "email_addr", "note")
    return {k: p.get(k) for k in keys if p.get(k) is not None}

def scor_similaritate_nume(a, b):
    return SequenceMatcher(None, normalizeaza_nume(a), normalizeaza_nume(b)).ratio()

PARENT_KIND_TO_EDGE = {"birth": "BIRTH_PARENT", "adoptive": "ADOPTIVE_PARENT", "step": "STEP_PARENT"}
CHILD_KIND_EXPR = "coalesce(c.kind, CASE WHEN coalesce(c.adopted, false) THEN 'adoptive' ELSE 'birth' END)"

def _normalizeaza_kind_filiatie(kind=None, adopted=None):
    if kind is not None and str(kind).strip():
        k = str(kind).strip().lower()
        if k in ("birth", "biological", "biologic", "biologica", "biologică", "bio", "natural"):
            return "birth"
        if k in ("adoptive", "adopted", "adoptat", "adoptata", "adoptată", "adoptiv", "adoptivă", "adopt"):
            return "adoptive"
        if k in ("step", "vitreg", "vitrega", "vitregă", "foster", "stepparent", "step_parent"):
            return "step"
    if adopted is not None:
        if isinstance(adopted, str):
            return "adoptive" if adopted.strip().lower() in ("true", "1", "yes", "da") else "birth"
        return "adoptive" if adopted else "birth"
    return None

def _seteaza_kind_child(s, uid: str, rid: str, cid: str, kind):
    if not kind:
        return
    s.run(
        "MATCH (rel:Relation {id:$rid, user_id:$uid})-[ch:CHILD]->(c:Person {id:$cid, user_id:$uid}) "
        "SET ch.kind = $kind, ch.adopted = $adopted",
        rid=rid, cid=cid, uid=uid, kind=kind, adopted=(kind == "adoptive"),
    )

def _ataseaza_parinti(s, uid: str, copil_id: str, father_id=None, mother_id=None, kind=None):
    fid = father_id.strip() if father_id and father_id.strip() else None
    mid = mother_id.strip() if mother_id and mother_id.strip() else None
    if not fid and not mid:
        return

    rid = None
    rel = s.run(
        "MATCH (rel:Relation {user_id:$uid})-[:CHILD]->(c:Person {id:$cid, user_id:$uid}) "
        "RETURN rel.id AS rid LIMIT 1",
        uid=uid, cid=copil_id,
    ).single()
    if rel:
        rid = rel["rid"]
    elif fid and mid:
        rel = s.run(
            "MATCH (rel:Relation {user_id:$uid})-[:MAN]->(:Person {id:$fid, user_id:$uid}) "
            "MATCH (rel)-[:WOMAN]->(:Person {id:$mid, user_id:$uid}) "
            "RETURN rel.id AS rid LIMIT 1",
            uid=uid, fid=fid, mid=mid,
        ).single()
        if rel:
            rid = rel["rid"]
    if not rid:
        rid = nou_id(uid, "r")
        s.run("CREATE (rel:Relation {id:$rid, user_id:$uid, type:$tip})",
              rid=rid, uid=uid, tip="unknown")

    s.run(
        "MATCH (rel:Relation {id:$rid, user_id:$uid}) "
        "MATCH (c:Person {id:$cid, user_id:$uid}) "
        "MERGE (rel)-[ch:CHILD]->(c) "
        "ON CREATE SET ch.seq = 0, ch.adopted = false",
        rid=rid, cid=copil_id, uid=uid,
    )
    _seteaza_kind_child(s, uid, rid, copil_id, _normalizeaza_kind_filiatie(kind=kind))
    if fid:
        s.run("MATCH (rel:Relation {id:$rid, user_id:$uid})-[r:MAN]->() DELETE r", rid=rid, uid=uid)
        s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) MATCH (m:Person {id:$fid, user_id:$uid}) "
              "CREATE (rel)-[:MAN]->(m)", rid=rid, fid=fid, uid=uid)
    if mid:
        s.run("MATCH (rel:Relation {id:$rid, user_id:$uid})-[r:WOMAN]->() DELETE r", rid=rid, uid=uid)
        s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) MATCH (w:Person {id:$mid, user_id:$uid}) "
              "CREATE (rel)-[:WOMAN]->(w)", rid=rid, mid=mid, uid=uid)

def _ataseaza_partener(s, uid: str, person_id: str, partner_id=None, partner_type=None):
    partener = partner_id.strip() if partner_id and partner_id.strip() else None
    if not partener or partener == person_id:
        return
    tip = _normalizeaza_tip_relatie_import(partner_type) if partner_type and partner_type.strip() else "married"

    existing = s.run(
        '\n        MATCH (rel:Relation {user_id:$uid})\n        WHERE ((rel)-[:MAN]->(:Person {id:$a, user_id:$uid}) AND (rel)-[:WOMAN]->(:Person {id:$b, user_id:$uid}))\n           OR ((rel)-[:MAN]->(:Person {id:$b, user_id:$uid}) AND (rel)-[:WOMAN]->(:Person {id:$a, user_id:$uid}))\n        RETURN rel.id AS rid LIMIT 1\n        ',
        uid=uid, a=person_id, b=partener,
    ).single()
    if existing:

        return

    rand = s.run("MATCH (p:Person {id:$pid, user_id:$uid}) RETURN p.gender AS g",
                 pid=person_id, uid=uid).single()
    gen = (rand["g"] or "").upper() if rand else ""
    man, woman = (person_id, partener) if gen == "M" else (partener, person_id)

    rid = nou_id(uid, "r")
    s.run("CREATE (rel:Relation {id:$rid, user_id:$uid, type:$tip})", rid=rid, uid=uid, tip=tip)
    s.run(
        "MATCH (rel:Relation {id:$rid, user_id:$uid}) "
        "MATCH (m:Person {id:$man, user_id:$uid}) MATCH (w:Person {id:$woman, user_id:$uid}) "
        "CREATE (rel)-[:MAN]->(m) CREATE (rel)-[:WOMAN]->(w)",
        rid=rid, man=man, woman=woman, uid=uid,
    )

def _harta_filiatie(s, uid: str) -> dict:
    copii_map = {}
    for row in s.run(
        '\n        MATCH (rel:Relation {user_id:$uid})-[:CHILD]->(c:Person {user_id:$uid})\n        MATCH (rel)-[:MAN|WOMAN]->(p:Person {user_id:$uid})\n        RETURN p.id AS p, c.id AS c\n        ',
        uid=uid,
    ):
        copii_map.setdefault(row["p"], []).append(row["c"])
    return copii_map

def _date_persoane(s, uid: str, ids: list) -> dict:
    ids = [i for i in ids if i]
    if not ids:
        return {}
    return {
        row["id"]: {"full_name": row["full_name"], "birth": row["birth"],
                    "death": row["death"], "gender": row["gender"]}
        for row in s.run(
            "MATCH (p:Person {user_id:$uid}) WHERE p.id IN $ids "
            "RETURN p.id AS id, p.full_name AS full_name, p.birth AS birth, "
            "       p.death AS death, p.gender AS gender",
            uid=uid, ids=ids,
        )
    }

def _e_confirmat(confirm) -> bool:
    return bool(confirm) and str(confirm).strip().lower() in ("true", "1", "yes", "da")

def _ridica_avertismente(avertismente: list, confirm):
    if avertismente and not _e_confirmat(confirm):
        raise HTTPException(409, {"warnings": avertismente})

def _valideaza_scriere_persoana(s, uid, pid, birth, death,
                                father_id=None, mother_id=None, partner_id=None,
                                parinti_relatie=None, confirm=None):
    erori, avertismente = [], []

    msg = eroare_deces_inainte_nastere(birth, death)
    if msg:
        erori.append(msg)

    fid = father_id.strip() if father_id and father_id.strip() else None
    mid = mother_id.strip() if mother_id and mother_id.strip() else None
    paid = partner_id.strip() if partner_id and partner_id.strip() else None

    if pid:
        if pid in (fid, mid):
            erori.append("O persoană nu poate fi propriul ei părinte.")
            fid = None if fid == pid else fid
            mid = None if mid == pid else mid
        if paid == pid:
            erori.append("O persoană nu poate fi propriul ei partener.")
            paid = None

    copii_map = None
    if pid and (fid or mid or paid):
        copii_map = _harta_filiatie(s, uid)

    if pid and copii_map is not None:
        for rol, par in (("tatăl", fid), ("mama", mid)):
            if par and este_descendent(copii_map, pid, par):
                erori.append(
                    f"Legătura ar crea un ciclu de filiație: {rol} ales(ă) este "
                    f"deja descendentul/descendenta persoanei."
                )

    if erori:
        raise HTTPException(422, " ".join(erori))

    parinti_ids = [x for x in (fid, mid) if x] + list(parinti_relatie or [])
    if birth is not None and parinti_ids:
        date = _date_persoane(s, uid, parinti_ids)
        for par_id in parinti_ids:
            if par_id in date:
                avertismente += avertismente_varste_parinte(date[par_id], birth)

    if pid and paid and copii_map is not None:
        rel = linie_directa(copii_map, pid, paid)
        if rel == "ascendent":
            avertismente.append("Partenerul ales este descendentul DIRECT al persoanei.")
        elif rel == "descendent":
            avertismente.append("Partenerul ales este ascendentul DIRECT al persoanei.")

    _ridica_avertismente(avertismente, confirm)

def _valideaza_partener_pereche(s, uid, id_a, id_b, confirm):
    a = id_a.strip() if id_a and id_a.strip() else None
    b = id_b.strip() if id_b and id_b.strip() else None
    if a and b and a == b:
        raise HTTPException(422, "O persoană nu poate fi propriul ei partener.")
    if not a or not b:
        return
    copii_map = _harta_filiatie(s, uid)
    rel = linie_directa(copii_map, a, b)
    if rel:
        _ridica_avertismente(
            ["Cei doi parteneri sunt în linie directă ascendent–descendent."], confirm
        )

app = FastAPI(title="Arbore Genealogic API")
app.add_middleware(GZipMiddleware, minimum_size=500)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

try:
    driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASS))
    driver.verify_connectivity()
    print(f"Neo4j conectat: {NEO4J_URI}")
except Exception as e:
    print(f"Eroare Neo4j: {e}")
    sys.exit(1)

@app.on_event("startup")
async def create_indexes():
    with driver.session() as s:
        s.run("CREATE INDEX person_id      IF NOT EXISTS FOR (p:Person)   ON (p.id)")
        s.run("CREATE INDEX person_user    IF NOT EXISTS FOR (p:Person)   ON (p.user_id)")
        s.run("CREATE INDEX person_id_user IF NOT EXISTS FOR (p:Person)   ON (p.id, p.user_id)")
        s.run("CREATE INDEX relation_id      IF NOT EXISTS FOR (r:Relation) ON (r.id)")
        s.run("CREATE INDEX relation_user    IF NOT EXISTS FOR (r:Relation) ON (r.user_id)")
        s.run("CREATE INDEX relation_id_user IF NOT EXISTS FOR (r:Relation) ON (r.id, r.user_id)")

        s.run("CREATE INDEX member_tree    IF NOT EXISTS FOR (m:Member)    ON (m.tree_id)")
        s.run("CREATE INDEX member_user    IF NOT EXISTS FOR (m:Member)    ON (m.user_id)")
        s.run("CREATE INDEX collab_token   IF NOT EXISTS FOR (c:CollabLink) ON (c.token)")
        s.run("CREATE INDEX collab_tree    IF NOT EXISTS FOR (c:CollabLink) ON (c.tree_id)")
        s.run("CREATE INDEX appuser_id     IF NOT EXISTS FOR (u:AppUser)   ON (u.user_id)")
        s.run("CREATE INDEX audit_tree     IF NOT EXISTS FOR (a:AuditLog)   ON (a.tree_id)")
        s.run("CREATE INDEX change_tree    IF NOT EXISTS FOR (c:ChangeRequest) ON (c.tree_id)")
        s.run("CREATE INDEX change_status  IF NOT EXISTS FOR (c:ChangeRequest) ON (c.status)")
    print("Indecsi Neo4j verificati/creati")

@app.get("/api/stats")
def statistici(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        nr_p = s.run("MATCH (p:Person   {user_id:$uid}) RETURN count(p) AS n", uid=uid).single()["n"]
        nr_r = s.run("MATCH (r:Relation {user_id:$uid}) RETURN count(r) AS n", uid=uid).single()["n"]
        rez = s.run('\n            MATCH (p:Person {user_id:$uid})\n            OPTIONAL MATCH path = (p)-[:BIRTH_PARENT|ADOPTIVE_PARENT*]->(desc:Person {user_id:$uid})\n            RETURN max(length(path)) AS max_gen\n        ', uid=uid).single()
        gen = rez["max_gen"] or 0
    return {"total_persons": nr_p, "total_relations": nr_r, "max_generation": gen}

@app.get("/api/data-quality")
def calitate_date(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        total = s.run("MATCH (p:Person {user_id:$uid}) RETURN count(p) AS n", uid=uid).single()["n"]
        fara_nastere = s.run(
            "MATCH (p:Person {user_id:$uid}) WHERE p.birth IS NULL RETURN count(p) AS n", uid=uid
        ).single()["n"]
        fara_gen = s.run(
            "MATCH (p:Person {user_id:$uid}) WHERE p.gender IS NULL OR p.gender='' RETURN count(p) AS n", uid=uid
        ).single()["n"]
        fara_poza = s.run(
            "MATCH (p:Person {user_id:$uid}) WHERE p.photo_url IS NULL AND p.photo IS NULL RETURN count(p) AS n", uid=uid
        ).single()["n"]
        fara_parinti = s.run(
            "\n            MATCH (p:Person {user_id:$uid})\n            WHERE NOT (:Relation {user_id:$uid})-[:CHILD]->(p)\n            RETURN count(p) AS n\n            ",
            uid=uid,
        ).single()["n"]
        fara_relatii = s.run(
            "\n            MATCH (p:Person {user_id:$uid})\n            WHERE NOT (:Relation {user_id:$uid})-[:CHILD]->(p)\n              AND NOT (:Relation {user_id:$uid})-[:MAN|WOMAN]->(p)\n            RETURN count(p) AS n\n            ",
            uid=uid,
        ).single()["n"]
        rel_incomplete = s.run(
            "\n            MATCH (r:Relation {user_id:$uid})\n            WHERE NOT (r)-[:MAN]->() OR NOT (r)-[:WOMAN]->()\n            RETURN count(r) AS n\n            ",
            uid=uid,
        ).single()["n"]
        audit_count = s.run(
            "MATCH (a:AuditLog {tree_id:$uid}) RETURN count(a) AS n", uid=uid
        ).single()["n"]
        pending_count = s.run(
            "MATCH (c:ChangeRequest {tree_id:$uid, status:'pending'}) RETURN count(c) AS n", uid=uid
        ).single()["n"]

    completate = 0
    total_campuri = max(total * 3, 1)
    completate += total - fara_nastere
    completate += total - fara_gen
    completate += total - fara_poza
    scor = round((completate / total_campuri) * 100)
    return {
        "score": scor,
        "total_persons": total,
        "missing_birth": fara_nastere,
        "missing_gender": fara_gen,
        "missing_photo": fara_poza,
        "without_parents": fara_parinti,
        "isolated_persons": fara_relatii,
        "incomplete_relations": rel_incomplete,
        "audit_entries": audit_count,
        "pending_changes": pending_count,
    }

@app.get("/api/tree")
def arbore(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        noduri = []
        for row in s.run("MATCH (p:Person {user_id:$uid}) RETURN p ORDER BY p.birth", uid=uid):
            noduri.append(serializeaza_persoana(row["p"]))

        partner_edges_raw = list(s.run(
            '\n            MATCH (rel:Relation {user_id:$uid})\n            MATCH (rel)-[:MAN]->(m:Person {user_id:$uid})\n            MATCH (rel)-[:WOMAN]->(f:Person {user_id:$uid})\n            RETURN m.id AS male_id, f.id AS female_id,\n                   rel.id AS relation_id, rel.type AS partner_type\n            ',
            uid=uid,
        ))
        perechi_vazute = set()
        partner_edges = []
        for row in partner_edges_raw:
            mid, fid = row["male_id"], row["female_id"]
            cheie = tuple(sorted([mid, fid]))
            if cheie not in perechi_vazute:
                perechi_vazute.add(cheie)
                partner_edges.append({
                    "source": mid, "target": fid, "type": "PARTNER",
                    "relation_id": row["relation_id"],
                    "partner_type": row["partner_type"],
                })

        parent_edges = [
            dict(row) for row in s.run(
                "\n                MATCH (rel:Relation {user_id:$uid})-[c:CHILD]->(copil:Person {user_id:$uid})\n                OPTIONAL MATCH (rel)-[:MAN]->(tata:Person {user_id:$uid})\n                OPTIONAL MATCH (rel)-[:WOMAN]->(mama:Person {user_id:$uid})\n                WITH rel, copil, tata, mama,\n                     CHILD_KIND AS kind, c.seq AS seq\n                WITH rel, copil, kind, seq,\n                     [p IN [tata, mama] WHERE p IS NOT NULL] AS parinti\n                UNWIND parinti AS parinte\n                RETURN parinte.id AS source, copil.id AS target,\n                       rel.id AS relation_id, seq AS seq,\n                       CASE kind WHEN 'adoptive' THEN 'ADOPTIVE_PARENT' WHEN 'step' THEN 'STEP_PARENT' ELSE 'BIRTH_PARENT' END AS type\n                ".replace("CHILD_KIND", CHILD_KIND_EXPR),
                uid=uid,
            )
        ]

    return {"nodes": noduri, "edges": partner_edges + parent_edges}

@app.delete("/api/tree")
def sterge_arbore(cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        nr = s.run("MATCH (p:Person {user_id:$uid}) RETURN count(p) AS n", uid=uid).single()["n"]
        s.run("MATCH (p:Person {user_id:$uid}) DETACH DELETE p", uid=uid)
        s.run("MATCH (r:Relation {user_id:$uid}) DETACH DELETE r", uid=uid)
        audit_log(s, cu, "delete_tree", "tree", uid, detalii={"persoane_sterse": nr})
    return {"status": "tree_deleted", "persoane_sterse": nr}

@app.get("/api/persons")
def lista_persoane(search: str = None, cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        if search:
            rows = s.run(
                "MATCH (p:Person {user_id:$uid}) WHERE toLower(p.full_name) CONTAINS toLower($s) RETURN p",
                uid=uid, s=search,
            )
        else:
            rows = s.run("MATCH (p:Person {user_id:$uid}) RETURN p ORDER BY p.birth", uid=uid)
        return [serializeaza_persoana(r["p"]) for r in rows]

@app.get("/api/persons/{pid}")
def detalii_persoana(pid: str, cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run("MATCH (p:Person {id:$pid, user_id:$uid}) RETURN p", pid=pid, uid=uid).single()
        if not row:
            raise HTTPException(404, "Persoana nu exista")
        persoana = serializeaza_persoana(row["p"])

        parteneri_rows = s.run(
            '\n            MATCH (p:Person {id:$pid, user_id:$uid})\n            MATCH (rel:Relation {user_id:$uid})\n            WHERE (rel)-[:MAN]->(p) OR (rel)-[:WOMAN]->(p)\n            OPTIONAL MATCH (rel)-[:MAN]->(m:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:WOMAN]->(f:Person {user_id:$uid})\n            WITH rel, CASE WHEN m.id = $pid THEN f ELSE m END AS partener\n            WHERE partener IS NOT NULL\n            RETURN DISTINCT partener, rel.type AS rel_type, rel.id AS rel_id\n            ',
            pid=pid, uid=uid,
        )
        parteneri = [
            {**serializeaza_persoana(r["partener"]), "rel_type": r["rel_type"], "rel_id": r["rel_id"]}
            for r in parteneri_rows
        ]

        copii_rows = s.run(
            ('\n            MATCH (p:Person {id:$pid, user_id:$uid})\n            MATCH (rel:Relation {user_id:$uid})\n            WHERE (rel)-[:MAN]->(p) OR (rel)-[:WOMAN]->(p)\n            MATCH (rel)-[c:CHILD]->(copil:Person {user_id:$uid})\n            RETURN copil, rel.id AS rel_id, CHILD_KIND AS kind, c.seq AS seq\n            ORDER BY seq\n            ').replace("CHILD_KIND", CHILD_KIND_EXPR),
            pid=pid, uid=uid,
        )
        copii_pe_tip = {"birth": [], "adoptive": [], "step": []}
        copii_vazuti = set()
        for r in copii_rows:
            d = serializeaza_persoana(r["copil"])
            d["rel_id"] = r["rel_id"]
            kind = r["kind"] or "birth"
            cheie = (d.get("id"), kind)
            if cheie in copii_vazuti:
                continue
            copii_vazuti.add(cheie)
            copii_pe_tip.setdefault(kind, []).append(d)

        parinti_rel_rows = s.run(
            ('\n            MATCH (p:Person {id:$pid, user_id:$uid})\n            MATCH (rel:Relation {user_id:$uid})-[c:CHILD]->(p)\n            OPTIONAL MATCH (rel)-[:MAN]->(tata:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:WOMAN]->(mama:Person {user_id:$uid})\n            RETURN rel.id AS rel_id, rel.type AS rel_type, CHILD_KIND AS kind, tata, mama\n            ').replace("CHILD_KIND", CHILD_KIND_EXPR),
            pid=pid, uid=uid,
        )
        parent_relations = []
        parinti_pe_tip = {"birth": [], "adoptive": [], "step": []}
        parinti_vazuti = set()
        for r in parinti_rel_rows:
            kind = r["kind"] or "birth"
            tata = serializeaza_persoana(r["tata"]) if r["tata"] else None
            mama = serializeaza_persoana(r["mama"]) if r["mama"] else None
            parent_relations.append({
                "rel_id": r["rel_id"], "rel_type": r["rel_type"], "kind": kind,
                "father": tata, "mother": mama,
            })
            for parinte in (tata, mama):
                if parinte is None:
                    continue
                cheie = (parinte.get("id"), kind)
                if cheie in parinti_vazuti:
                    continue
                parinti_vazuti.add(cheie)
                parinti_pe_tip.setdefault(kind, []).append(parinte)

    return {
        "person":   persoana,
        "spouses":  parteneri,
        "children": {"biological": copii_pe_tip["birth"], "step": copii_pe_tip["step"], "adopted": copii_pe_tip["adoptive"]},
        "parents":  {"biological": parinti_pe_tip["birth"], "step": parinti_pe_tip["step"], "adopted": parinti_pe_tip["adoptive"]},
        "parent_relations": parent_relations,
    }

@app.get("/api/relations")
def lista_relatii(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        rows = s.run(
            '\n            MATCH (rel:Relation {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:MAN]->(m:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:WOMAN]->(f:Person {user_id:$uid})\n            RETURN rel, m, f ORDER BY rel.id\n            ',
            uid=uid,
        )
        rezultat = []
        for r in rows:
            rel_dict = dict(r["rel"])
            rel_dict["male"]   = serializeaza_persoana(r["m"]) if r["m"] else None
            rel_dict["female"] = serializeaza_persoana(r["f"]) if r["f"] else None
            rezultat.append(rel_dict)
    return rezultat

@app.get("/api/relations/{rel_id}")
def detalii_relatie(rel_id: str, cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run(
            '\n            MATCH (rel:Relation {id:$rid, user_id:$uid})\n            OPTIONAL MATCH (rel)-[:MAN]->(m:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:WOMAN]->(f:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[c:CHILD]->(copil:Person {user_id:$uid})\n            RETURN rel, m, f, collect({person: copil, seq: c.seq, adopted: c.adopted}) AS copii\n            ',
            rid=rel_id, uid=uid,
        ).single()
    if not row:
        raise HTTPException(404, "Relatia nu exista")
    rel_dict = dict(row["rel"])
    rel_dict["male"]     = serializeaza_persoana(row["m"]) if row["m"] else None
    rel_dict["female"]   = serializeaza_persoana(row["f"]) if row["f"] else None
    rel_dict["children"] = [
        {**serializeaza_persoana(c["person"]), "seq": c["seq"], "adopted": c["adopted"]}
        for c in row["copii"] if c["person"] is not None
    ]
    return rel_dict

@app.post("/api/persons")
async def creaza_persoana(
    full_name:    str  = Form(...),
    given_name:   str  = Form(None),
    surname:      str  = Form(None),
    gender:       str  = Form(None),
    birth:        str  = Form(None),
    death:        str  = Form(None),
    note:         str  = Form(None),
    tel:          str  = Form(None),
    email:        str  = Form(None),
    address:      str  = Form(None),
    relation_id:  str  = Form(None),
    father_id:    str  = Form(None),
    mother_id:    str  = Form(None),
    parent_kind:  str  = Form(None),
    partner_id:   str  = Form(None),
    partner_type: str  = Form("married"),
    photo_url:    str  = Form(None),
    confirm:      str  = Form(None),
    cu:           dict = Depends(utilizator_curent),
):
    cere_scriere(cu)
    uid = cu["user_id"]
    pid = nou_id(uid, "p")

    props = {"id": pid, "user_id": uid, "full_name": full_name.strip()}
    if given_name and given_name.strip(): props["given_name"] = given_name.strip()
    if surname    and surname.strip():    props["surname"]    = surname.strip()
    if gender     and gender.strip():     props["gender"]     = gender.strip().upper()
    bi = int_sau_none(birth); di = int_sau_none(death)
    if bi is not None: props["birth"] = bi
    if di is not None: props["death"] = di
    if note    and note.strip():    props["note"]       = note.strip()
    if tel     and tel.strip():     props["tel"]        = tel.strip()
    if email   and email.strip():   props["email_addr"] = email.strip()
    if address and address.strip(): props["address"]    = address.strip()
    if photo_url and photo_url.strip(): props["photo_url"] = photo_url.strip()

    with driver.session() as s:

        parinti_rel = []
        if relation_id and relation_id.strip():
            parinti_rel = [r["pid"] for r in s.run(
                "MATCH (rel:Relation {id:$rid, user_id:$uid})-[:MAN|WOMAN]->(p:Person {user_id:$uid}) "
                "RETURN p.id AS pid",
                rid=relation_id.strip(), uid=uid,
            )]
        _valideaza_scriere_persoana(
            s, uid, None, bi, di,
            father_id=father_id, mother_id=mother_id, partner_id=partner_id,
            parinti_relatie=parinti_rel, confirm=confirm,
        )

        set_clause = ", ".join(f"p.{k} = ${k}" for k in props)
        s.run(f"CREATE (p:Person) SET {set_clause}", **props)

        if relation_id and relation_id.strip():
            s.run(
                '\n                MATCH (rel:Relation {id:$rid, user_id:$uid})\n                MATCH (copil:Person {id:$pid, user_id:$uid})\n                MERGE (rel)-[ch:CHILD]->(copil)\n                ON CREATE SET ch.seq = 0, ch.adopted = false\n                ',
                rid=relation_id.strip(), pid=pid, uid=uid,
            )
            _seteaza_kind_child(s, uid, relation_id.strip(), pid, _normalizeaza_kind_filiatie(kind=parent_kind))
        elif father_id or mother_id:
            _ataseaza_parinti(s, uid, pid, father_id, mother_id, kind=parent_kind)

        _ataseaza_partener(s, uid, pid, partner_id, partner_type)
        audit_log(s, cu, "create", "person", pid, dupa=props)

    return {"status": "created", "id": pid}

@app.put("/api/persons/{pid}")
async def editeaza_persoana(
    pid:          str,
    full_name:    str  = Form(None),
    given_name:   str  = Form(None),
    surname:      str  = Form(None),
    gender:       str  = Form(None),
    birth:        str  = Form(None),
    death:        str  = Form(None),
    note:         str  = Form(None),
    tel:          str  = Form(None),
    email:        str  = Form(None),
    address:      str  = Form(None),
    photo_url:    str  = Form(None),
    father_id:    str  = Form(None),
    mother_id:    str  = Form(None),
    partner_id:   str  = Form(None),
    partner_type: str  = Form(None),
    adopted:      str  = Form(None),
    parent_kind:  str  = Form(None),
    confirm:      str  = Form(None),
    cu:           dict = Depends(utilizator_curent),
):
    cere_scriere(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run("MATCH (p:Person {id:$p, user_id:$u}) RETURN p", p=pid, u=uid).single()
        if not row:
            raise HTTPException(404, "Persoana nu exista")
        existent = dict(row["p"])

        birth_final = int_sau_none(birth) if birth is not None else existent.get("birth")
        death_final = int_sau_none(death) if death is not None else existent.get("death")
        _valideaza_scriere_persoana(
            s, uid, pid, birth_final, death_final,
            father_id=father_id, mother_id=mother_id, partner_id=partner_id,
            confirm=confirm,
        )

    seturi, stergeri = [], []
    params = {"pid": pid, "uid": uid}

    def add(key, val):
        if val is None:
            stergeri.append(f"p.{key}")
        else:
            seturi.append(f"p.{key} = ${key}")
            params[key] = val

    if full_name  and full_name.strip():  add("full_name",  full_name.strip())
    if given_name is not None:            add("given_name", given_name.strip())
    if surname    is not None:            add("surname",    surname.strip())
    if gender     and gender.strip():     add("gender",     gender.strip().upper())
    if birth is not None: add("birth", int_sau_none(birth))
    if death is not None: add("death", int_sau_none(death))
    if note    is not None: add("note",       note.strip()    or None)
    if tel     is not None: add("tel",        tel.strip()     or None)
    if email   is not None: add("email_addr", email.strip()   or None)
    if address is not None: add("address",    address.strip() or None)
    if photo_url is not None and photo_url.strip(): add("photo_url", photo_url.strip())

    with driver.session() as s:
        if seturi or stergeri:
            cypher = "MATCH (p:Person {id:$pid, user_id:$uid})"
            if seturi:   cypher += " SET "    + ", ".join(seturi)
            if stergeri: cypher += " REMOVE " + ", ".join(stergeri)
            s.run(cypher, **params)

        kind = _normalizeaza_kind_filiatie(kind=parent_kind, adopted=adopted)
        _ataseaza_parinti(s, uid, pid, father_id, mother_id, kind=kind)
        _ataseaza_partener(s, uid, pid, partner_id, partner_type)
        dupa = s.run("MATCH (p:Person {id:$pid, user_id:$uid}) RETURN p", pid=pid, uid=uid).single()
        audit_log(
            s, cu, "update", "person", pid,
            inainte=mini_persoana(existent),
            dupa=mini_persoana(dict(dupa["p"]) if dupa else {}),
        )

    return {"status": "updated"}

@app.delete("/api/persons/{pid}")
def sterge_persoana(pid: str, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run("MATCH (p:Person {id:$p, user_id:$u}) RETURN p", p=pid, u=uid).single()
        if not row:
            raise HTTPException(404, "Persoana nu exista")
        inainte = mini_persoana(dict(row["p"]))
        s.run("MATCH (p:Person {id:$p, user_id:$u}) DETACH DELETE p", p=pid, u=uid)
        s.run(
            '\n            MATCH (rel:Relation {user_id:$uid})\n            WHERE NOT (rel)-[:MAN]->() AND NOT (rel)-[:WOMAN]->() AND NOT (rel)-[:CHILD]->()\n            DELETE rel\n            ',
            uid=uid,
        )
        audit_log(s, cu, "delete", "person", pid, inainte=inainte)
    return {"status": "deleted"}

class PhotoMatch(BaseModel):
    filename: str
    url: str

class PhotoMatchRequest(BaseModel):
    photos: List[PhotoMatch]

class ChangeRequestCreate(BaseModel):
    reason: str = ""

def serializeaza_cerere(row):
    c = dict(row["c"])
    for key in ("payload_json", "result_json"):
        if c.get(key):
            try:
                c[key.replace("_json", "")] = json.loads(c[key])
            except Exception:
                c[key.replace("_json", "")] = {}
    return c

@app.post("/api/change-requests/person-delete/{pid}")
def cere_stergere_persoana(pid: str, req: ChangeRequestCreate, cu: dict = Depends(utilizator_curent)):
    cere_scriere(cu)
    if cu.get("role") == "owner":
        raise HTTPException(400, "Proprietarul poate sterge direct persoana.")
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run("MATCH (p:Person {id:$pid, user_id:$uid}) RETURN p", pid=pid, uid=uid).single()
        if not row:
            raise HTTPException(404, "Persoana nu exista")
        existent = s.run(
            "MATCH (c:ChangeRequest {tree_id:$uid, entity_id:$pid, action:'person_delete', status:'pending'}) "
            "RETURN c LIMIT 1",
            uid=uid, pid=pid,
        ).single()
        if existent:
            return {"status": "already_pending", "id": dict(existent["c"]).get("id")}
        cid = nou_id(uid, "chg")
        payload = {"person": mini_persoana(dict(row["p"])), "reason": req.reason or ""}
        s.run(
            "CREATE (c:ChangeRequest {id:$id, tree_id:$uid, requester_id:$rid, requester_email:$email, "
            "requester_name:$name, action:'person_delete', entity_type:'person', entity_id:$pid, "
            "status:'pending', reason:$reason, payload_json:$payload, created_at:$created})",
            id=cid, uid=uid, rid=cu.get("caller_id"), email=cu.get("email") or "",
            name=cu.get("full_name") or "", pid=pid, reason=req.reason or "",
            payload=json_dump(payload), created=acum_iso(),
        )
        audit_log(s, cu, "request_delete", "person", pid, detalii={"change_request_id": cid})
    return {"status": "pending", "id": cid}

@app.get("/api/change-requests")
def listeaza_cereri(status: str = "pending", cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        rows = s.run(
            "MATCH (c:ChangeRequest {tree_id:$uid}) "
            "WHERE $status = 'all' OR c.status = $status "
            "RETURN c ORDER BY c.created_at DESC LIMIT 100",
            uid=cu["user_id"], status=status,
        )
        return {"items": [serializeaza_cerere(r) for r in rows]}

@app.get("/api/audit")
def lista_audit(limit: int = 50, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    limit = max(1, min(int(limit or 50), 200))
    with driver.session() as s:
        rows = s.run(
            "MATCH (a:AuditLog {tree_id:$uid}) RETURN a ORDER BY a.created_at DESC LIMIT $limit",
            uid=cu["user_id"], limit=limit,
        )
        out = []
        for r in rows:
            a = dict(r["a"])
            for key in ("before_json", "after_json", "details_json"):
                if a.get(key):
                    try:
                        a[key.replace("_json", "")] = json.loads(a[key])
                    except Exception:
                        a[key.replace("_json", "")] = {}
            out.append(a)
        return {"items": out}

@app.post("/api/change-requests/{change_id}/reject")
def respinge_cerere(change_id: str, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        row = s.run(
            "MATCH (c:ChangeRequest {id:$id, tree_id:$uid, status:'pending'}) RETURN c",
            id=change_id, uid=cu["user_id"],
        ).single()
        if not row:
            raise HTTPException(404, "Cererea nu exista sau nu mai este in asteptare.")
        s.run(
            "MATCH (c:ChangeRequest {id:$id, tree_id:$uid}) "
            "SET c.status='rejected', c.reviewed_at=$now, c.reviewer_id=$rid",
            id=change_id, uid=cu["user_id"], now=acum_iso(), rid=cu.get("caller_id"),
        )
        audit_log(s, cu, "reject_change", "change_request", change_id)
    return {"status": "rejected"}

@app.post("/api/change-requests/{change_id}/approve")
def aproba_cerere(change_id: str, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run(
            "MATCH (c:ChangeRequest {id:$id, tree_id:$uid, status:'pending'}) RETURN c",
            id=change_id, uid=uid,
        ).single()
        if not row:
            raise HTTPException(404, "Cererea nu exista sau nu mai este in asteptare.")
        cerere = dict(row["c"])
        if cerere.get("action") != "person_delete":
            raise HTTPException(422, "Tip de cerere nesuportat.")
        pid = cerere.get("entity_id")
        prow = s.run("MATCH (p:Person {id:$pid, user_id:$uid}) RETURN p", pid=pid, uid=uid).single()
        if not prow:
            s.run(
                "MATCH (c:ChangeRequest {id:$id, tree_id:$uid}) "
                "SET c.status='applied', c.reviewed_at=$now, c.reviewer_id=$rid, "
                "c.result_json=$result",
                id=change_id, uid=uid, now=acum_iso(), rid=cu.get("caller_id"),
                result=json_dump({"status": "person_already_missing"}),
            )
            return {"status": "applied", "detail": "person_already_missing"}
        inainte = mini_persoana(dict(prow["p"]))
        s.run("MATCH (p:Person {id:$pid, user_id:$uid}) DETACH DELETE p", pid=pid, uid=uid)
        s.run(
            '\n            MATCH (rel:Relation {user_id:$uid})\n            WHERE NOT (rel)-[:MAN]->() AND NOT (rel)-[:WOMAN]->() AND NOT (rel)-[:CHILD]->()\n            DELETE rel\n            ',
            uid=uid,
        )
        s.run(
            "MATCH (c:ChangeRequest {id:$id, tree_id:$uid}) "
            "SET c.status='applied', c.reviewed_at=$now, c.reviewer_id=$rid, c.result_json=$result",
            id=change_id, uid=uid, now=acum_iso(), rid=cu.get("caller_id"),
            result=json_dump({"deleted_person_id": pid}),
        )
        audit_log(s, cu, "approve_change", "change_request", change_id, detalii={"deleted_person_id": pid})
        audit_log(s, cu, "delete", "person", pid, inainte=inainte,
                  detalii={"approved_change_request_id": change_id})
    return {"status": "applied", "deleted_person_id": pid}

@app.post("/api/persons/match-photos")
def potriveste_poze(req: PhotoMatchRequest, cu: dict = Depends(utilizator_curent)):
    cere_scriere(cu)
    uid = cu["user_id"]

    harta = {}
    for p in req.photos:
        nume = p.filename.strip()
        harta[nume.lower()] = p.url
        baza = os.path.splitext(nume)[0].lower()
        harta.setdefault(baza, p.url)

    actualizate = 0
    with driver.session() as s:
        rows = list(s.run(
            "MATCH (p:Person {user_id:$uid}) WHERE p.photo IS NOT NULL "
            "RETURN p.id AS id, p.photo AS photo",
            uid=uid,
        ))
        nume_in_db = set()
        for r in rows:
            nume_foto = (r["photo"] or "").strip().lower()
            if not nume_foto:
                continue
            nume_in_db.add(nume_foto)
            nume_in_db.add(os.path.splitext(nume_foto)[0])
            baza_foto = os.path.splitext(nume_foto)[0]
            url = harta.get(nume_foto) or harta.get(baza_foto)
            if url:
                s.run(
                    "MATCH (p:Person {id:$pid, user_id:$uid}) SET p.photo_url = $url",
                    pid=r["id"], uid=uid, url=url,
                )
                actualizate += 1

    nepotrivite = []
    for p in req.photos:
        nume = p.filename.strip().lower()
        baza = os.path.splitext(nume)[0]
        if nume not in nume_in_db and baza not in nume_in_db:
            nepotrivite.append(p.filename)

    return {
        "status": "matched",
        "persoane_actualizate": actualizate,
        "poze_nepotrivite": nepotrivite,
    }

@app.post("/api/persons/normalize-photo-names")
def normalizeaza_nume_poze(cu: dict = Depends(utilizator_curent)):
    cere_scriere(cu)
    uid = cu["user_id"]
    renames = {}
    actualizate = 0
    with driver.session() as s:
        rows = list(s.run(
            "MATCH (p:Person {user_id:$uid}) WHERE p.photo IS NOT NULL "
            "RETURN p.id AS id, p.photo AS photo",
            uid=uid,
        ))
        for r in rows:
            old = (r["photo"] or "").strip()
            if not old:
                continue
            new = slug_nume_fisier(old)
            if new == old:
                continue
            props = {"photo": new}
            if SUPABASE_URL:
                props["photo_url"] = (
                    f"{SUPABASE_URL}/storage/v1/object/public"
                    f"/{SUPABASE_BUCKET}/{uid}/{new}"
                )
            s.run("MATCH (p:Person {id:$pid, user_id:$uid}) SET p += $props",
                  pid=r["id"], uid=uid, props=props)
            actualizate += 1
            renames[old] = new
    return {
        "status": "normalized",
        "persoane_actualizate": actualizate,
        "renames": [{"old": k, "new": v} for k, v in renames.items()],
    }

@app.post("/api/relations")
async def creaza_relatie(
    male_id:       str  = Form(None),
    female_id:     str  = Form(None),
    relation_type: str  = Form("married"),
    confirm:       str  = Form(None),
    cu:            dict = Depends(utilizator_curent),
):
    cere_scriere(cu)
    uid = cu["user_id"]
    rid = nou_id(uid, "r")
    with driver.session() as s:

        _valideaza_partener_pereche(s, uid, male_id, female_id, confirm)
        s.run("CREATE (rel:Relation {id:$rid, user_id:$uid, type:$tip})",
              rid=rid, uid=uid, tip=_normalizeaza_tip_relatie_import(relation_type))
        if male_id and male_id.strip():
            s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) MATCH (m:Person {id:$mid, user_id:$uid}) CREATE (rel)-[:MAN]->(m)",
                  rid=rid, mid=male_id.strip(), uid=uid)
        if female_id and female_id.strip():
            s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) MATCH (f:Person {id:$fid, user_id:$uid}) CREATE (rel)-[:WOMAN]->(f)",
                  rid=rid, fid=female_id.strip(), uid=uid)
        audit_log(
            s, cu, "create", "relation", rid,
            dupa={"id": rid, "male_id": male_id, "female_id": female_id, "type": relation_type or "married"},
        )
    return {"status": "created", "id": rid}

@app.put("/api/relations/{rel_id}")
async def editeaza_relatie(
    rel_id:        str,
    male_id:       str  = Form(None),
    female_id:     str  = Form(None),
    relation_type: str  = Form(None),
    confirm:       str  = Form(None),
    cu:            dict = Depends(utilizator_curent),
):
    cere_scriere(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        rel_initial = s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) RETURN rel", rid=rel_id, uid=uid).single()
        if not rel_initial:
            raise HTTPException(404, "Relatia nu exista")
        inainte = dict(rel_initial["rel"])

        if male_id is not None or female_id is not None:
            cur = s.run(
                '\n                MATCH (rel:Relation {id:$rid, user_id:$uid})\n                OPTIONAL MATCH (rel)-[:MAN]->(m:Person {user_id:$uid})\n                OPTIONAL MATCH (rel)-[:WOMAN]->(f:Person {user_id:$uid})\n                RETURN m.id AS mid, f.id AS fid\n                ',
                rid=rel_id, uid=uid,
            ).single()
            mid_final = male_id   if male_id   is not None else (cur["mid"] or "")
            fid_final = female_id if female_id is not None else (cur["fid"] or "")
            _valideaza_partener_pereche(s, uid, mid_final, fid_final, confirm)

        if relation_type:
            s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) SET rel.type = $tip",
                  rid=rel_id, uid=uid, tip=relation_type)
        if male_id is not None:
            s.run("MATCH (rel:Relation {id:$rid, user_id:$uid})-[r:MAN]->() DELETE r", rid=rel_id, uid=uid)
            if male_id.strip():
                s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) MATCH (m:Person {id:$mid, user_id:$uid}) CREATE (rel)-[:MAN]->(m)",
                      rid=rel_id, mid=male_id.strip(), uid=uid)
        if female_id is not None:
            s.run("MATCH (rel:Relation {id:$rid, user_id:$uid})-[r:WOMAN]->() DELETE r", rid=rel_id, uid=uid)
            if female_id.strip():
                s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) MATCH (f:Person {id:$fid, user_id:$uid}) CREATE (rel)-[:WOMAN]->(f)",
                      rid=rel_id, fid=female_id.strip(), uid=uid)
        rel_final = s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) RETURN rel", rid=rel_id, uid=uid).single()
        audit_log(s, cu, "update", "relation", rel_id, inainte=inainte,
                  dupa=dict(rel_final["rel"]) if rel_final else {})
    return {"status": "updated"}

@app.post("/api/relations/{rel_id}/children")
async def adauga_copil_la_relatie(
    rel_id:   str,
    child_id: str  = Form(...),
    seq:      int  = Form(0),
    adopted:  str  = Form("false"),
    kind:     str  = Form(None),
    confirm:  str  = Form(None),
    cu:       dict = Depends(utilizator_curent),
):
    cere_scriere(cu)
    uid = cu["user_id"]
    kind_final = _normalizeaza_kind_filiatie(kind=kind, adopted=adopted) or "birth"
    este_adoptat = kind_final == "adoptive"
    cid = child_id.strip()
    with driver.session() as s:
        if not s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) RETURN rel", rid=rel_id, uid=uid).single():
            raise HTTPException(404, "Relatia nu exista")

        parinti = [r["pid"] for r in s.run(
            "MATCH (rel:Relation {id:$rid, user_id:$uid})-[:MAN|WOMAN]->(p:Person {user_id:$uid}) "
            "RETURN p.id AS pid",
            rid=rel_id, uid=uid,
        )]
        if cid in parinti:
            raise HTTPException(422, "O persoană nu poate fi propriul ei părinte.")
        if parinti:
            copii_map = _harta_filiatie(s, uid)
            for par in parinti:
                if este_descendent(copii_map, cid, par):
                    raise HTTPException(
                        422,
                        "Legătura ar crea un ciclu de filiație: unul dintre părinții "
                        "relației este deja descendentul copilului.",
                    )
            date = _date_persoane(s, uid, parinti + [cid])
            copil_birth = (date.get(cid) or {}).get("birth")
            avertismente = []
            if copil_birth is not None:
                for par in parinti:
                    if par in date:
                        avertismente += avertismente_varste_parinte(date[par], copil_birth)
            _ridica_avertismente(avertismente, confirm)

        s.run(
            '\n            MATCH (rel:Relation {id:$rid, user_id:$uid})\n            MATCH (copil:Person {id:$cid, user_id:$uid})\n            MERGE (rel)-[c:CHILD]->(copil)\n            SET c.seq = $seq, c.adopted = $adopted, c.kind = $kind\n            ',
            rid=rel_id, cid=child_id.strip(), uid=uid, seq=seq, adopted=este_adoptat, kind=kind_final,
        )
        audit_log(s, cu, "add_child", "relation", rel_id,
                  detalii={"child_id": cid, "kind": kind_final, "adopted": este_adoptat, "seq": seq})
    return {"status": "child_added"}

@app.delete("/api/relations/{rel_id}/children/{child_id}")
def elimina_copil_din_relatie(rel_id: str, child_id: str, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        s.run(
            "MATCH (rel:Relation {id:$rid, user_id:$uid})-[c:CHILD]->(p:Person {id:$cid, user_id:$uid}) DELETE c",
            rid=rel_id, cid=child_id, uid=uid,
        )
        audit_log(s, cu, "remove_child", "relation", rel_id, detalii={"child_id": child_id})
    return {"status": "child_removed"}

@app.delete("/api/relations/{rel_id}")
def sterge_relatie(rel_id: str, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        row = s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) RETURN rel", rid=rel_id, uid=uid).single()
        if not row:
            raise HTTPException(404, "Relatia nu exista")
        inainte = dict(row["rel"])
        s.run("MATCH (rel:Relation {id:$rid, user_id:$uid}) DETACH DELETE rel", rid=rel_id, uid=uid)
        audit_log(s, cu, "delete", "relation", rel_id, inainte=inainte)
    return {"status": "deleted"}

@app.get("/api/integrity")
def verifica_integritate(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    probleme = []

    with driver.session() as s:
        persoane = {
            row["id"]: {"full_name": row["full_name"], "birth": row["birth"],
                        "death": row["death"], "gender": row["gender"]}
            for row in s.run(
                "MATCH (p:Person {user_id:$uid}) RETURN p.id AS id, p.full_name AS full_name, "
                "p.birth AS birth, p.death AS death, p.gender AS gender",
                uid=uid,
            )
        }

        parinte_copil = [
            (row["p"], row["c"], row["rol"]) for row in s.run(
                '\n                MATCH (rel:Relation {user_id:$uid})-[:CHILD]->(c:Person {user_id:$uid})\n                MATCH (rel)-[r:MAN|WOMAN]->(p:Person {user_id:$uid})\n                RETURN p.id AS p, c.id AS c, type(r) AS rol\n                ',
                uid=uid,
            )
        ]
        parteneri = [
            (row["a"], row["b"]) for row in s.run(
                '\n                MATCH (rel:Relation {user_id:$uid})\n                MATCH (rel)-[:MAN]->(a:Person {user_id:$uid})\n                MATCH (rel)-[:WOMAN]->(b:Person {user_id:$uid})\n                RETURN a.id AS a, b.id AS b\n                ',
                uid=uid,
            )
        ]

    nume = lambda pid: (persoane.get(pid) or {}).get("full_name") or pid

    for pid, p in persoane.items():
        msg = eroare_deces_inainte_nastere(p.get("birth"), p.get("death"))
        if msg:
            probleme.append({
                "tip": "deces_inainte_nastere", "severitate": "eroare",
                "person_ids": [pid], "mesaj_ro": f"{nume(pid)}: {msg}",
            })

    for p, c, _rol in parinte_copil:
        if p == c:
            probleme.append({
                "tip": "propriul_parinte", "severitate": "eroare",
                "person_ids": [p], "mesaj_ro": f"{nume(p)} apare ca propriul ei părinte.",
            })
    for a, b in parteneri:
        if a == b:
            probleme.append({
                "tip": "propriul_partener", "severitate": "eroare",
                "person_ids": [a], "mesaj_ro": f"{nume(a)} apare ca propriul ei partener.",
            })

    copii_map = {}
    for p, c, _rol in parinte_copil:
        if p != c:
            copii_map.setdefault(p, []).append(c)
    for ciclu in gaseste_cicluri(copii_map):
        probleme.append({
            "tip": "ciclu_filiatie", "severitate": "eroare", "person_ids": ciclu,
            "mesaj_ro": "Ciclu de filiație: " + " → ".join(nume(x) for x in ciclu)
                        + f" → {nume(ciclu[0])}.",
        })

    vazute = set()
    for p, c, rol in parinte_copil:
        if (p, c) in vazute or p == c:
            continue
        vazute.add((p, c))
        par = dict(persoane.get(p) or {})

        if rol == "WOMAN":
            par["gender"] = "F"
        for msg in avertismente_varste_parinte(par, (persoane.get(c) or {}).get("birth")):
            probleme.append({
                "tip": "varsta_parinte", "severitate": "avertisment",
                "person_ids": [p, c],
                "mesaj_ro": f"{msg} (copil: {nume(c)})",
            })

    for a, b in parteneri:
        if a != b and linie_directa(copii_map, a, b):
            probleme.append({
                "tip": "partener_linie_directa", "severitate": "avertisment",
                "person_ids": [a, b],
                "mesaj_ro": f"{nume(a)} și {nume(b)} sunt parteneri, dar și în linie "
                            f"directă ascendent–descendent.",
            })

    parinti_dupa_copil = {}
    for p, c, rol in parinte_copil:
        pp = persoane.get(p) or {}
        cc = persoane.get(c) or {}
        parinti_dupa_copil.setdefault(c, {}).setdefault(rol, set()).add(p)
        if pp.get("death") is not None and cc.get("birth") is not None:
            limita = pp["death"] + (1 if rol == "MAN" else 0)
            if cc["birth"] > limita:
                probleme.append({
                    "tip": "copil_dupa_deces_parinte", "severitate": "avertisment",
                    "person_ids": [p, c],
                    "mesaj_ro": f"{nume(c)} este nascut dupa decesul parintelui {nume(p)}.",
                })
        if rol == "MAN" and (pp.get("gender") or "").upper() == "F":
            probleme.append({
                "tip": "gen_rol_parinte", "severitate": "avertisment",
                "person_ids": [p, c],
                "mesaj_ro": f"{nume(p)} este marcata femeie, dar apare ca tata pentru {nume(c)}.",
            })
        if rol == "WOMAN" and (pp.get("gender") or "").upper() == "M":
            probleme.append({
                "tip": "gen_rol_parinte", "severitate": "avertisment",
                "person_ids": [p, c],
                "mesaj_ro": f"{nume(p)} este marcat barbat, dar apare ca mama pentru {nume(c)}.",
            })

    for copil, roluri in parinti_dupa_copil.items():
        if len(roluri.get("MAN", set())) > 1:
            probleme.append({
                "tip": "doi_tati_biologici", "severitate": "eroare",
                "person_ids": [copil, *roluri["MAN"]],
                "mesaj_ro": f"{nume(copil)} are mai multi tati biologici.",
            })
        if len(roluri.get("WOMAN", set())) > 1:
            probleme.append({
                "tip": "doua_mame_biologice", "severitate": "eroare",
                "person_ids": [copil, *roluri["WOMAN"]],
                "mesaj_ro": f"{nume(copil)} are mai multe mame biologice.",
            })

    for pid, p in persoane.items():
        lipsuri = []
        if not p.get("birth"):
            lipsuri.append("an nastere")
        if not p.get("gender"):
            lipsuri.append("gen")
        if lipsuri:
            probleme.append({
                "tip": "date_incomplete", "severitate": "info",
                "person_ids": [pid],
                "mesaj_ro": f"{nume(pid)} are date incomplete: {', '.join(lipsuri)}.",
            })

    return {"probleme": probleme, "total": len(probleme)}

@app.get("/api/duplicates")
def detecteaza_duplicate(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        persoane = [serializeaza_persoana(row["p"]) for row in s.run(
            "MATCH (p:Person {user_id:$uid}) RETURN p", uid=uid,
        )]
        parinti = {
            row["id"]: set(row["parents"]) for row in s.run(
                "\n                MATCH (p:Person {user_id:$uid})\n                OPTIONAL MATCH (rel:Relation {user_id:$uid})-[:CHILD]->(p)\n                OPTIONAL MATCH (rel)-[:MAN|WOMAN]->(par:Person {user_id:$uid})\n                RETURN p.id AS id, collect(DISTINCT par.id) AS parents\n                ",
                uid=uid,
            )
        }

    campuri = ("id", "full_name", "birth", "death", "gender", "photo_url", "address")
    mini = lambda p: {k: p.get(k) for k in campuri}
    pairs = []
    for i in range(len(persoane)):
        for j in range(i + 1, len(persoane)):
            a, b = persoane[i], persoane[j]
            nume_score = scor_similaritate_nume(a.get("full_name"), b.get("full_name"))
            if nume_score < 0.72:
                continue
            score = int(nume_score * 55)
            motive = [f"nume similar {int(nume_score * 100)}%"]
            ba, bb = a.get("birth"), b.get("birth")
            if ba is None or bb is None:
                score += 10
                motive.append("an nastere lipsa la una dintre persoane")
            else:
                diff = abs(int(ba) - int(bb))
                if diff == 0:
                    score += 25
                    motive.append("acelasi an de nastere")
                elif diff <= 2:
                    score += 15
                    motive.append("ani de nastere apropiati")
                elif diff <= 5:
                    score += 6
                    motive.append("ani de nastere relativ apropiati")
                else:
                    score -= 20
            if a.get("gender") and b.get("gender") and a.get("gender") == b.get("gender"):
                score += 8
                motive.append("acelasi gen")
            comuni = parinti.get(a.get("id"), set()) & parinti.get(b.get("id"), set())
            if comuni:
                score += 20
                motive.append("parinti comuni")
            score = max(0, min(100, score))
            if score >= 60:
                pairs.append({"a": mini(a), "b": mini(b), "score": score, "motive": motive})
    pairs.sort(key=lambda x: x["score"], reverse=True)
    return {
        "pairs": pairs
    }

class MergeRequest(BaseModel):
    keep_id: str
    remove_id: str

@app.post("/api/persons/merge")
def uneste_duplicate(req: MergeRequest, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    keep, rm = req.keep_id.strip(), req.remove_id.strip()
    if not keep or not rm or keep == rm:
        raise HTTPException(422, "Alege două persoane diferite pentru unire.")

    with driver.session() as s:
        with s.begin_transaction() as tx:
            rk = tx.run("MATCH (p:Person {id:$id, user_id:$uid}) RETURN p", id=keep, uid=uid).single()
            rr = tx.run("MATCH (p:Person {id:$id, user_id:$uid}) RETURN p", id=rm, uid=uid).single()
            if not rk or not rr:
                raise HTTPException(404, "Una dintre persoane nu există.")
            keep_props = dict(rk["p"])
            rm_props   = dict(rr["p"])

            for rol in ("MAN", "WOMAN"):
                tx.run(
                    f"\n                    MATCH (k:Person {{id:$keep, user_id:$uid}})\n                    MATCH (rel:Relation {{user_id:$uid}})-[:{rol}]->(d:Person {{id:$rm, user_id:$uid}})\n                    WHERE NOT (rel)-[:MAN|WOMAN]->(k)\n                    MERGE (rel)-[:{rol}]->(k)\n                    ",
                    keep=keep, rm=rm, uid=uid,
                )

            tx.run(
                '\n                MATCH (k:Person {id:$keep, user_id:$uid})\n                MATCH (rel:Relation {user_id:$uid})-[x:CHILD]->(d:Person {id:$rm, user_id:$uid})\n                WHERE NOT (rel)-[:CHILD]->(k) AND NOT (rel)-[:MAN|WOMAN]->(k)\n                MERGE (rel)-[nc:CHILD]->(k)\n                SET nc.seq = coalesce(x.seq, 0), nc.adopted = coalesce(x.adopted, false)\n                ',
                keep=keep, rm=rm, uid=uid,
            )

            de_completat = {}
            for k, v in rm_props.items():
                if k in ("id", "user_id"):
                    continue
                gol = keep_props.get(k) is None or str(keep_props.get(k)).strip() == ""
                if gol and v is not None and str(v).strip() != "":
                    de_completat[k] = v
            if de_completat:
                set_clause = ", ".join(f"p.{k} = ${k}" for k in de_completat)
                tx.run(
                    f"MATCH (p:Person {{id:$keep, user_id:$uid}}) SET {set_clause}",
                    keep=keep, uid=uid, **de_completat,
                )

            tx.run("MATCH (p:Person {id:$rm, user_id:$uid}) DETACH DELETE p", rm=rm, uid=uid)
            tx.run(
                '\n                MATCH (rel:Relation {user_id:$uid})\n                WHERE NOT (rel)-[:MAN]->() AND NOT (rel)-[:WOMAN]->() AND NOT (rel)-[:CHILD]->()\n                DELETE rel\n                ',
                uid=uid,
            )
            tx.commit()

        audit_log(
            s, cu, "merge", "person", keep,
            inainte={"keep": mini_persoana(keep_props), "remove": mini_persoana(rm_props)},
            detalii={"removed_id": rm},
        )

    return {"status": "merged", "keep_id": keep, "removed_id": rm}

@app.get("/api/relationship")
def calculator_rudenie(
    from_id: str = Query(..., alias="from"),
    to_id:   str = Query(..., alias="to"),
    cu:      dict = Depends(utilizator_curent),
):
    uid = cu["user_id"]

    if from_id == to_id:
        with driver.session() as s:
            row = s.run("MATCH (p:Person {id:$id, user_id:$uid}) RETURN p",
                        id=from_id, uid=uid).single()
        if not row:
            raise HTTPException(404, f"Persoana {from_id} nu există în acest arbore")
        p = serializeaza_persoana(row["p"])
        return {
            "found": True, "label": "aceeași persoană",
            "chain": [{"person_id": p["id"], "full_name": p.get("full_name"), "step_ro": ""}],
            "path_ids": [p["id"]],
        }

    with driver.session() as s:
        for test_id in (from_id, to_id):
            if not s.run("MATCH (p:Person {id:$id, user_id:$uid}) RETURN p",
                         id=test_id, uid=uid).single():
                raise HTTPException(404, f"Persoana {test_id} nu există în acest arbore")

        rows = list(s.run(
            "\n            MATCH (a:Person {id:$a, user_id:$uid}), (b:Person {id:$b, user_id:$uid})\n            MATCH p = allShortestPaths((a)-[:MAN|WOMAN|CHILD*..40]-(b))\n            WHERE all(n IN nodes(p) WHERE n.user_id = $uid)\n            RETURN [n IN nodes(p) | {id: n.id, full_name: n.full_name, gender: n.gender,\n                                     este_relatie: 'Relation' IN labels(n)}] AS noduri,\n                   [r IN relationships(p) | {type: type(r),\n                                             adopted: coalesce(r.adopted, false)}] AS muchii\n            LIMIT 50\n            ",
            a=from_id, b=to_id, uid=uid,
        ))

    candidati = []
    for r in rows:
        pasi = decodeaza_drum(r["noduri"], r["muchii"])
        if pasi is not None:
            candidati.append((r["noduri"], r["muchii"], pasi))

    if not candidati:
        return {
            "found": False,
            "label": "Nu există o legătură de rudenie înregistrată între cele două persoane",
            "chain": [], "path_ids": [],
        }

    def cheie(c):
        noduri, _, pasi = c
        return (numara_alianta(pasi),
                sum(1 for p in pasi if p["adopted"]),
                [str(n.get("id")) for n in noduri])
    noduri, muchii, pasi = min(candidati, key=cheie)

    return {"found": True, **construieste_raspuns(noduri, muchii, pasi)}

@app.get("/api/export/template")
def descarca_sablon(cu: dict = Depends(utilizator_curent)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "Libraria openpyxl lipseste")

    wb = openpyxl.Workbook()

    culori = {"Persons": "4472C4", "Relations": "70AD47", "Children": "17BECF"}

    ws_g = wb.active
    ws_g.title = "Ghid"
    ghid = [
        ("📖 Ce poți scrie în coloane", True),
        ("", False),
        ("Gender (Persons):  M  sau  F", False),
        ("Photo (Persons):  numele fișierului, ex: ion.jpg", False),
        ("Type (Relations):  married · divorced · separated · partner · engaged   (gol = married)", False),
        ("Adopted (Children):  true  sau  false", False),
    ]
    for text, bold in ghid:
        ws_g.append([text])
        if bold:
            ws_g.cell(row=ws_g.max_row, column=1).font = Font(bold=True)
    ws_g.column_dimensions["A"].width = 100

    ws_p = wb.create_sheet("Persons")
    ws_p.append(["ID", "Full name", "Given name", "Surname at birth", "Gender",
                 "Birth year", "Death year", "Note", "Tel", "Email", "Address", "Photo"])
    ws_p.append(["P1", "Ion Popescu", "Ion", "Popescu", "M", 1950, None,
                 "Exemplu notă", "0722000000", "ion@exemplu.com", "Brașov", "ion_popescu.jpg"])
    ws_p.append(["P2", "Maria Ionescu", "Maria", "Ionescu", "F", 1952, None,
                 "", "", "", "București", "maria_ionescu.jpg"])
    ws_p.append(["P3", "Andrei Popescu", "Andrei", "Popescu", "M", 1978, None,
                 "", "", "", "", "andrei_popescu.jpg"])

    ws_r = wb.create_sheet("Relations")
    ws_r.append(["Relation ID", "Male ID", "Female ID", "Type"])
    ws_r.append(["R1", "P1", "P2", "married"])

    ws_c = wb.create_sheet("Children")
    ws_c.append(["Relation ID", "Child ID", "Adopted"])
    ws_c.append(["R1", "P3", "false"])

    for sheet_name, culoare in culori.items():
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=culoare)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sablon_arbore_genealogic.xlsx"},
    )

def _citeste_sheet(wb, nume_posibile: list):
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in [n.lower() for n in nume_posibile]:
            return wb[sheet_name]
    return None

def _headers(ws) -> dict:
    return {str(c.value or "").strip().lower(): i for i, c in enumerate(ws[1])}

def _val(row, cm: dict, cheie: str):
    idx = cm.get(cheie)
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    return str(v).strip() if v is not None else None

def _normalizeaza_tip_relatie_import(tip):
    val = str(tip or "married").strip().lower()
    if not val:
        return "married"
    if val in ("married", "marr", "casatorit", "casatorita", "căsătorit", "căsătorită", "current", "y"):
        return "married"
    if val in ("divorced", "div", "divortat", "divortata", "divorțat", "divorțată", "former"):
        return "divorced"
    if val in ("partner", "partener", "partenera", "parteneră", "relationship", "concubinaj",
               "unmarried", "necasatorit", "necasatorita", "necăsătorit", "necăsătorită"):
        return "partner"
    if val in ("engaged", "logodit", "logodita", "logodită", "logodna", "logodnă", "fiance", "fiancee"):
        return "engaged"
    if val in ("separated", "separat", "separata", "separată", "despartit", "despărțit"):
        return "separated"
    return val

IMPORT_EXECUTOR = ThreadPoolExecutor(max_workers=int(os.getenv("IMPORT_WORKERS", "2")))
IMPORT_BATCH_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", "1000"))
IMPORT_JOBS = {}
IMPORT_LOCK = threading.Lock()

def _chunks(rows, size=IMPORT_BATCH_SIZE):
    for i in range(0, len(rows), size):
        yield rows[i:i + size]

def _import_now_iso():
    return datetime.now(timezone.utc).isoformat()

def _job_public(job: dict):
    return {
        "job_id": job["job_id"],
        "type": job["type"],
        "status": job["status"],
        "stage": job.get("stage"),
        "progress": job.get("progress", 0),
        "message": job.get("message", ""),
        "result": job.get("result"),
        "error": job.get("error"),
        "created_at": job.get("created_at"),
        "updated_at": job.get("updated_at"),
    }

def _set_import_job(job_id: str, **updates):
    with IMPORT_LOCK:
        job = IMPORT_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = _import_now_iso()

def _salveaza_upload_import(file: UploadFile, suffix: str) -> str:
    os.makedirs(tempfile.gettempdir(), exist_ok=True)
    safe_suffix = suffix if suffix.startswith(".") else f".{suffix}"
    tmp = os.path.join(tempfile.gettempdir(), f"import_{uuid.uuid4().hex}{safe_suffix}")
    with open(tmp, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return tmp

def _ruleaza_import_job(job_id: str, kind: str, tmp: str, cu: dict):
    def update(stage=None, progress=None, message=None):
        data = {}
        if stage is not None:
            data["stage"] = stage
        if progress is not None:
            data["progress"] = progress
        if message is not None:
            data["message"] = message
        if data:
            _set_import_job(job_id, **data)

    _set_import_job(job_id, status="running", stage="start", progress=1, message="Import pornit")
    try:
        if kind == "excel":
            result = _proceseaza_import_excel(tmp, cu, update)
        else:
            result = _proceseaza_import_gedcom(tmp, cu, update)
        _set_import_job(
            job_id,
            status="done",
            stage="done",
            progress=100,
            message="Import finalizat",
            result=result,
        )
    except HTTPException as e:
        _set_import_job(job_id, status="failed", stage="failed", progress=100, error=e.detail)
    except Exception as e:
        detalii = traceback.format_exc()
        print(f"\nEROARE IMPORT JOB {job_id}:\n{detalii}")
        _set_import_job(job_id, status="failed", stage="failed", progress=100, error=str(e))

def _porneste_import_job(kind: str, tmp: str, cu: dict):
    job_id = uuid.uuid4().hex
    job = {
        "job_id": job_id,
        "type": kind,
        "user_id": cu["user_id"],
        "caller_id": cu.get("caller_id"),
        "status": "queued",
        "stage": "queued",
        "progress": 0,
        "message": "Import in coada",
        "created_at": _import_now_iso(),
        "updated_at": _import_now_iso(),
    }
    with IMPORT_LOCK:
        IMPORT_JOBS[job_id] = job
    IMPORT_EXECUTOR.submit(_ruleaza_import_job, job_id, kind, tmp, dict(cu))
    return _job_public(job)

def _proceseaza_import_excel(tmp: str, cu: dict, update=None):
    uid = cu["user_id"]
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "Libraria openpyxl lipseste")

    try:
        if update:
            update("parse", 5, "Citesc fisierul Excel")
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

        ws_p = _citeste_sheet(wb, ["persons", "persoane", "sheet1", "oameni"])
        if ws_p is None:
            raise HTTPException(400, "Lipseste foaia 'Persons' din Excel")

        cm_p = _headers(ws_p)
        persoane = []
        for row in ws_p.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            pid_orig = _val(row, cm_p, "id")
            full_name = _val(row, cm_p, "full name") or _val(row, cm_p, "full_name") or _val(row, cm_p, "name")
            if not full_name and not pid_orig:
                continue
            persoane.append({
                "id": pid_orig, "full_name": full_name or "",
                "given_name": _val(row, cm_p, "given name") or _val(row, cm_p, "given_name"),
                "surname": _val(row, cm_p, "surname") or _val(row, cm_p, "surname at birth"),
                "gender": (_val(row, cm_p, "gender") or _val(row, cm_p, "gen") or "").upper()[:1],
                "birth": int_sau_none(_val(row, cm_p, "birth") or _val(row, cm_p, "birth year")),
                "death": int_sau_none(_val(row, cm_p, "death") or _val(row, cm_p, "death year")),
                "note": _val(row, cm_p, "note") or _val(row, cm_p, "notes"),
                "tel": _val(row, cm_p, "tel") or _val(row, cm_p, "phone"),
                "email": _val(row, cm_p, "email"),
                "address": _val(row, cm_p, "address") or _val(row, cm_p, "adresa"),
                "photo": _val(row, cm_p, "photo"),
            })

        ws_r = _citeste_sheet(wb, ["relations", "relatii", "cupluri", "families"])
        relatii = []
        if ws_r:
            cm_r = _headers(ws_r)
            for row in ws_r.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                rid_orig = _val(row, cm_r, "relation id") or _val(row, cm_r, "id")
                if not rid_orig:
                    continue
                relatii.append({
                    "id": rid_orig,
                    "male_id": _val(row, cm_r, "male id") or _val(row, cm_r, "father id") or _val(row, cm_r, "man id"),
                    "female_id": _val(row, cm_r, "female id") or _val(row, cm_r, "mother id") or _val(row, cm_r, "woman id"),
                    "type": _normalizeaza_tip_relatie_import(_val(row, cm_r, "type")),
                })

        ws_c = _citeste_sheet(wb, ["children", "copii", "child"])
        copii_rel = []
        if ws_c:
            cm_c = _headers(ws_c)
            for row in ws_c.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                rid_orig = _val(row, cm_c, "relation id")
                cid_orig = _val(row, cm_c, "child id")
                if not rid_orig or not cid_orig:
                    continue
                copii_rel.append({
                    "relation_id": rid_orig, "child_id": cid_orig,
                    "seq": int_sau_none(_val(row, cm_c, "seq") or _val(row, cm_c, "order")) or 0,
                    "adopted": (_val(row, cm_c, "adopted") or "").lower() in ("yes", "da", "true", "1"),
                })

        wb.close()
        if update:
            update("write", 35, f"Scriu {len(persoane)} persoane in baza de date")

        nr_p_create = nr_p_update = nr_r_create = nr_r_update = nr_c = 0
        id_map_p, id_map_r = {}, {}

        with driver.session() as s:

            prefix_p = f"{uid}_p_"
            prefix_r = f"{uid}_r_"
            for row in s.run(
                "MATCH (p:Person {user_id:$uid}) WHERE p.id STARTS WITH $pfx RETURN p.id AS pid",
                uid=uid, pfx=prefix_p,
            ):
                id_map_p.setdefault(row["pid"][len(prefix_p):], row["pid"])
            for row in s.run(
                "MATCH (r:Relation {user_id:$uid}) WHERE r.id STARTS WITH $pfx RETURN r.id AS rid",
                uid=uid, pfx=prefix_r,
            ):
                id_map_r.setdefault(row["rid"][len(prefix_r):], row["rid"])

            persoane_batch = []
            for p in persoane:
                pid = genereaza_id_stabil(uid, p["id"], p["full_name"], p["birth"])
                id_map_p[str(p["id"]).strip() if p["id"] is not None else pid] = pid
                props = {"original_id": str(p["id"] or "")}
                if p["full_name"]:         props["full_name"]  = p["full_name"]
                if p["given_name"]:        props["given_name"] = p["given_name"]
                if p["surname"]:           props["surname"]    = p["surname"]
                if p["gender"]:            props["gender"]     = p["gender"]
                if p["birth"] is not None: props["birth"]      = p["birth"]
                if p["death"] is not None: props["death"]      = p["death"]
                if p["note"]:              props["note"]       = p["note"]
                if p["tel"]:               props["tel"]        = p["tel"]
                if p["email"]:             props["email_addr"] = p["email"]
                if p["address"]:           props["address"]    = p["address"]
                if p["photo"]:
                    nume_poza = slug_nume_fisier(p["photo"])
                    props["photo"] = nume_poza
                    if SUPABASE_URL:
                        props["photo_url"] = (
                            f"{SUPABASE_URL}/storage/v1/object/public"
                            f"/{SUPABASE_BUCKET}/{uid}/{nume_poza}"
                        )
                persoane_batch.append({"id": pid, "props": props})

            if persoane_batch:
                for batch in _chunks(persoane_batch):
                    rez = s.run(
                        '\n                    UNWIND $batch AS row\n                    MERGE (p:Person {id: row.id, user_id: $uid})\n                    ON CREATE SET p += row.props, p._nou = true\n                    ON MATCH  SET p += row.props, p._nou = false\n                    RETURN sum(CASE WHEN p._nou THEN 1 ELSE 0 END) AS created,\n                           sum(CASE WHEN p._nou THEN 0 ELSE 1 END) AS updated\n                    ',
                        uid=uid, batch=batch,
                    ).single()
                    if rez:
                        nr_p_create += rez["created"] or 0
                        nr_p_update += rez["updated"] or 0

                s.run("MATCH (p:Person {user_id:$uid}) REMOVE p._nou", uid=uid)
            if update:
                update("write", 55, "Scriu relatiile de cuplu")

            relatii_batch = []
            for rel in relatii:
                rid = genereaza_id_relatie_stabil(uid, rel["id"])
                id_map_r[str(rel["id"]).strip()] = rid

                male_pid   = id_map_p.get(str(rel["male_id"]).strip())   if rel["male_id"]   else None
                female_pid = id_map_p.get(str(rel["female_id"]).strip()) if rel["female_id"] else None
                relatii_batch.append({
                    "id": rid, "type": _normalizeaza_tip_relatie_import(rel["type"]),
                    "male_id": male_pid, "female_id": female_pid,
                })

            if relatii_batch:
                for batch in _chunks(relatii_batch):
                    rez = s.run(
                        '\n                    UNWIND $batch AS row\n                    MERGE (r:Relation {id: row.id, user_id: $uid})\n                    ON CREATE SET r.type = row.type, r._nou = true\n                    ON MATCH  SET r.type = row.type, r._nou = false\n                    RETURN sum(CASE WHEN r._nou THEN 1 ELSE 0 END) AS created,\n                           sum(CASE WHEN r._nou THEN 0 ELSE 1 END) AS updated\n                    ',
                        uid=uid, batch=batch,
                    ).single()
                    if rez:
                        nr_r_create += rez["created"] or 0
                        nr_r_update += rez["updated"] or 0
                s.run("MATCH (r:Relation {user_id:$uid}) REMOVE r._nou", uid=uid)

                rel_ids = [r["id"] for r in relatii_batch]
                for ids in _chunks(rel_ids):
                    s.run(
                        "UNWIND $ids AS rid MATCH (r:Relation {id:rid,user_id:$uid})-[x:MAN|WOMAN|CHILD]->() DELETE x",
                        ids=ids, uid=uid,
                    )
                man_batch = [{"rid": r["id"], "mid": r["male_id"]}
                             for r in relatii_batch if r["male_id"]]
                if man_batch:
                    for batch in _chunks(man_batch):
                        s.run(
                            '\n                        UNWIND $batch AS row\n                        MATCH (r:Relation {id:row.rid, user_id:$uid})\n                        MATCH (m:Person   {id:row.mid, user_id:$uid})\n                        CREATE (r)-[:MAN]->(m)\n                        ',
                            batch=batch, uid=uid,
                        )
                woman_batch = [{"rid": r["id"], "fid": r["female_id"]}
                               for r in relatii_batch if r["female_id"]]
                if woman_batch:
                    for batch in _chunks(woman_batch):
                        s.run(
                            '\n                        UNWIND $batch AS row\n                        MATCH (r:Relation {id:row.rid, user_id:$uid})\n                        MATCH (f:Person   {id:row.fid, user_id:$uid})\n                        CREATE (r)-[:WOMAN]->(f)\n                        ',
                            batch=batch, uid=uid,
                        )

            copii_batch = []
            if update:
                update("write", 75, "Leg copiii de familii")
            for c in copii_rel:
                rid = id_map_r.get(str(c["relation_id"]).strip())
                cid = id_map_p.get(str(c["child_id"]).strip())
                if rid and cid:
                    copii_batch.append({
                        "rid": rid, "cid": cid,
                        "seq": c["seq"], "adopted": c["adopted"],
                        "kind": "adoptive" if c["adopted"] else "birth",
                    })
            nr_c = len(copii_batch)

            if copii_batch:
                for batch in _chunks(copii_batch):
                    s.run(
                        '\n                    UNWIND $batch AS row\n                    MATCH (r:Relation {id:row.rid, user_id:$uid})\n                    MATCH (copil:Person {id:row.cid, user_id:$uid})\n                    MERGE (r)-[ch:CHILD]->(copil)\n                    SET ch.seq = row.seq, ch.adopted = row.adopted,\n                        ch.kind = coalesce(row.kind, CASE WHEN row.adopted THEN \'adoptive\' ELSE \'birth\' END)\n                    ',
                        uid=uid, batch=batch,
                    )

        rezultat = {
            "status": "imported",
            "persons_created": nr_p_create, "persons_updated": nr_p_update,
            "relations_created": nr_r_create, "relations_updated": nr_r_update,
            "children_linked": nr_c, "total_persons": len(persoane),
        }
        if update:
            update("audit", 95, "Finalizez importul")
        with driver.session() as audit_s:
            audit_log(audit_s, cu, "import_excel", "tree", uid, detalii=rezultat)
        return rezultat

    except HTTPException:
        raise
    except Exception as e:
        detalii = traceback.format_exc()
        print(f"\nEROARE IMPORT EXCEL:\n{detalii}")
        raise HTTPException(500, f"Eroare la import Excel: {e} | Detalii: {detalii[-500:]}")
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except PermissionError:
            pass

@app.post("/api/import/excel")
async def import_excel(file: UploadFile = File(...), cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    tmp = _salveaza_upload_import(file, ".xlsx")
    return _proceseaza_import_excel(tmp, cu)

@app.post("/api/import/excel/job")
async def import_excel_job(file: UploadFile = File(...), cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    tmp = _salveaza_upload_import(file, ".xlsx")
    return _porneste_import_job("excel", tmp, cu)

@app.get("/api/import/jobs/{job_id}")
def status_import_job(job_id: str, cu: dict = Depends(utilizator_curent)):
    with IMPORT_LOCK:
        job = IMPORT_JOBS.get(job_id)
        if not job:
            raise HTTPException(404, "Job de import negasit")
        if job.get("user_id") != cu["user_id"]:
            raise HTTPException(403, "Nu ai acces la acest job de import")
        return _job_public(job)

@app.get("/api/export/gedcom")
def export_gedcom(cu: dict = Depends(utilizator_curent)):
    uid = cu["user_id"]
    with driver.session() as s:
        persoane = [dict(r["p"]) for r in s.run(
            "MATCH (p:Person {user_id:$uid}) RETURN p ORDER BY p.id", uid=uid,
        )]
        relatii = []
        for r in s.run(
            ('\n            MATCH (rel:Relation {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:MAN]->(m:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[:WOMAN]->(w:Person {user_id:$uid})\n            OPTIONAL MATCH (rel)-[c:CHILD]->(k:Person {user_id:$uid})\n            WITH rel, m, w, k, c ORDER BY coalesce(c.seq, 0), k.id\n            RETURN rel.id AS id, rel.type AS type,\n                   m.id AS male_id, w.id AS female_id,\n                   collect(CASE WHEN k IS NULL THEN NULL\n                           ELSE {id: k.id, adopted: coalesce(c.adopted, false), kind: CHILD_KIND} END) AS children\n            ORDER BY rel.id\n            ').replace("CHILD_KIND", CHILD_KIND_EXPR),
            uid=uid,
        ):
            d = dict(r)
            d["children"] = [c for c in d["children"] if c is not None]

            if d["male_id"] or d["female_id"] or d["children"]:
                relatii.append(d)

    text = genereaza_gedcom(persoane, relatii)
    return StreamingResponse(
        io.BytesIO(text.encode("utf-8")),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="arbore.ged"'},
    )

def _proceseaza_import_gedcom(tmp: str, cu: dict, update=None):
    uid = cu["user_id"]

    try:
        if update:
            update("parse", 5, "Citesc fisierul GEDCOM")
        for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
            try:
                with open(tmp, "r", encoding=enc, errors="strict") as f:
                    linii = f.readlines()
                break
            except UnicodeDecodeError:
                continue
        else:
            with open(tmp, "r", encoding="utf-8", errors="replace") as f:
                linii = f.readlines()

        indivizi, familii = parseaza_gedcom(linii)
        if not indivizi:
            raise HTTPException(400, "Fisierul GEDCOM nu contine persoane (INDI)")
        if update:
            update("transform", 25, f"Am gasit {len(indivizi)} persoane si {len(familii)} familii")

        nr_p_create = nr_p_update = nr_r_create = nr_r_update = nr_c = 0

        persoane_batch = []
        for ged_id, date in indivizi.items():
            pid = genereaza_id_stabil(uid, original_id=ged_id)
            props = {"original_id": ged_id}
            for k in ("full_name", "given_name", "surname", "gender", "birth", "death", "note", "tel", "email_addr", "address"):
                if k in date:
                    props[k] = date[k]
            if date.get("photo"):
                nume_poza = slug_nume_fisier(date["photo"])
                props["photo"] = nume_poza
                if SUPABASE_URL:
                    props["photo_url"] = (
                        f"{SUPABASE_URL}/storage/v1/object/public"
                        f"/{SUPABASE_BUCKET}/{uid}/{nume_poza}"
                    )
            persoane_batch.append({"id": pid, "props": props})

        relatii_batch = []
        copii_batch = []
        for fam_id, fam in familii.items():
            rid = genereaza_id_relatie_stabil(uid, fam_id)
            sot_ged = fam.get("sot")
            sotie_ged = fam.get("sotie")
            relatii_batch.append({
                "id": rid,

                "type": _normalizeaza_tip_relatie_import(fam.get("type", "married")),
                "male_id":   genereaza_id_stabil(uid, original_id=sot_ged)   if sot_ged in indivizi else None,
                "female_id": genereaza_id_stabil(uid, original_id=sotie_ged) if sotie_ged in indivizi else None,
            })
            for seq_i, copil_ged_id in enumerate(fam.get("copii", []), start=1):
                if copil_ged_id not in indivizi:
                    continue
                info_copil = indivizi.get(copil_ged_id, {})
                if fam_id in (info_copil.get("famc_adoptate") or []):
                    kind_copil = "adoptive"
                elif fam_id in (info_copil.get("famc_vitrege") or []):
                    kind_copil = "step"
                else:
                    kind_copil = "birth"
                copii_batch.append({
                    "rid": rid,
                    "cid": genereaza_id_stabil(uid, original_id=copil_ged_id),
                    "seq": seq_i,
                    "kind": kind_copil,
                    "adopted": kind_copil == "adoptive",
                })

        with driver.session() as s:
            if update:
                update("write", 40, "Scriu persoanele in baza de date")
            ids_p = [x["id"] for x in persoane_batch]
            existente_p = set()
            if ids_p:
                for ids in _chunks(ids_p):
                    existente_p.update(r["id"] for r in s.run(
                        "MATCH (p:Person {user_id:$uid}) WHERE p.id IN $ids RETURN p.id AS id",
                        uid=uid, ids=ids,
                    ))
            nr_p_update = len(existente_p)
            nr_p_create = len(ids_p) - nr_p_update
            if persoane_batch:
                for batch in _chunks(persoane_batch):
                    s.run(
                        '\n                    UNWIND $batch AS row\n                    MERGE (p:Person {id: row.id, user_id: $uid})\n                    SET p += row.props\n                    ',
                        uid=uid, batch=batch,
                    )

            ids_r = [x["id"] for x in relatii_batch]
            existente_r = set()
            if ids_r:
                for ids in _chunks(ids_r):
                    existente_r.update(r["id"] for r in s.run(
                        "MATCH (r:Relation {user_id:$uid}) WHERE r.id IN $ids RETURN r.id AS id",
                        uid=uid, ids=ids,
                    ))
            nr_r_update = len(existente_r)
            nr_r_create = len(ids_r) - nr_r_update
            if relatii_batch:
                if update:
                    update("write", 65, "Scriu familiile si cuplurile")

                for batch in _chunks(relatii_batch):
                    s.run(
                        '\n                    UNWIND $batch AS row\n                    MERGE (r:Relation {id: row.id, user_id: $uid})\n                    SET r.type = row.type\n                    WITH r\n                    OPTIONAL MATCH (r)-[x:MAN|WOMAN|CHILD]->() DELETE x\n                    ',
                        uid=uid, batch=batch,
                    )

                man_batch = [{"rid": r["id"], "mid": r["male_id"]}
                             for r in relatii_batch if r["male_id"]]
                if man_batch:
                    for batch in _chunks(man_batch):
                        s.run(
                            '\n                    UNWIND $batch AS row\n                    MATCH (r:Relation {id: row.rid, user_id: $uid})\n                    MATCH (m:Person {id: row.mid, user_id: $uid})\n                    CREATE (r)-[:MAN]->(m)\n                    ',
                            uid=uid, batch=batch,
                        )

                woman_batch = [{"rid": r["id"], "fid": r["female_id"]}
                               for r in relatii_batch if r["female_id"]]
                if woman_batch:
                    for batch in _chunks(woman_batch):
                        s.run(
                            '\n                    UNWIND $batch AS row\n                    MATCH (r:Relation {id: row.rid, user_id: $uid})\n                    MATCH (f:Person {id: row.fid, user_id: $uid})\n                    CREATE (r)-[:WOMAN]->(f)\n                    ',
                            uid=uid, batch=batch,
                        )

            if copii_batch:
                if update:
                    update("write", 82, "Leg copiii de familii")
                for batch in _chunks(copii_batch):
                    s.run(
                        '\n                    UNWIND $batch AS row\n                    MATCH (r:Relation {id: row.rid, user_id: $uid})\n                    MATCH (copil:Person {id: row.cid, user_id: $uid})\n                    MERGE (r)-[ch:CHILD]->(copil)\n                    SET ch.seq = row.seq, ch.adopted = row.adopted,\n                        ch.kind = coalesce(row.kind, CASE WHEN row.adopted THEN \'adoptive\' ELSE \'birth\' END)\n                    ',
                        uid=uid, batch=batch,
                    )
                nr_c = len(copii_batch)

            s.run(
                "\n                MATCH (p:Person {user_id:$uid})\n                WHERE p.original_id IS NULL AND p.full_name IS NULL AND NOT (p)--()\n                DELETE p\n                ",
                uid=uid,
            )

        rezultat = {
            "status": "imported",
            "persons_created": nr_p_create, "persons_updated": nr_p_update,
            "relations_created": nr_r_create, "relations_updated": nr_r_update,
            "children_linked": nr_c, "total_persons": len(indivizi),
        }
        if update:
            update("audit", 95, "Finalizez importul")
        with driver.session() as audit_s:
            audit_log(audit_s, cu, "import_gedcom", "tree", uid, detalii=rezultat)
        return rezultat

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Eroare la import GEDCOM: {e}")
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except PermissionError:
            pass

@app.post("/api/import/gedcom")
async def import_gedcom(file: UploadFile = File(...), cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    tmp = _salveaza_upload_import(file, ".ged")
    return _proceseaza_import_gedcom(tmp, cu)

@app.post("/api/import/gedcom/job")
async def import_gedcom_job(file: UploadFile = File(...), cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    tmp = _salveaza_upload_import(file, ".ged")
    return _porneste_import_job("gedcom", tmp, cu)

@app.get("/api/share")
def info_link_public(request: Request, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        r = s.run(
            "MATCH (sl:ShareLink {user_id:$uid}) RETURN sl.token AS t", uid=cu["user_id"]
        ).single()
    if not r or not r["t"]:
        return {"token": None, "url": None}
    return {"token": r["t"], "url": f"{frontend_url(request)}/view/{r['t']}"}

@app.post("/api/share/generate")
def genereaza_link_public(request: Request, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    token = secrets.token_urlsafe(16)
    with driver.session() as s:
        s.run("MERGE (sl:ShareLink {user_id:$uid}) SET sl.token = $token", uid=uid, token=token)
        audit_log(s, cu, "create_public_link", "share_link", uid)
    return {"token": token, "url": f"{frontend_url(request)}/view/{token}"}

@app.delete("/api/share/revoke")
def revocare_link_public(cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    uid = cu["user_id"]
    with driver.session() as s:
        s.run("MATCH (sl:ShareLink {user_id:$uid}) DELETE sl", uid=uid)
        audit_log(s, cu, "revoke_public_link", "share_link", uid)
    return {"message": "Link public dezactivat"}

@app.get("/api/share/view/{token}")
def vizualizeaza_arbore_public(token: str):
    with driver.session() as s:
        r = s.run("MATCH (sl:ShareLink {token:$t}) RETURN sl.user_id AS uid", t=token).single()
        if not r:
            raise HTTPException(404, "Link invalid sau expirat")
        uid = r["uid"]

        noduri = s.run("MATCH (p:Person {user_id:$uid}) RETURN p", uid=uid).data()

        partner_rows = list(s.run(
            '\n            MATCH (r:Relation {user_id:$uid})\n            MATCH (r)-[:MAN]->(m:Person {user_id:$uid})\n            MATCH (r)-[:WOMAN]->(w:Person {user_id:$uid})\n            RETURN m.id AS male_id, w.id AS female_id,\n                   r.id AS relation_id, r.type AS partner_type\n            ',
            uid=uid,
        ))

        parent_rows = list(s.run(
            "\n            MATCH (r:Relation {user_id:$uid})-[c:CHILD]->(copil:Person {user_id:$uid})\n            OPTIONAL MATCH (r)-[:MAN]->(tata:Person {user_id:$uid})\n            OPTIONAL MATCH (r)-[:WOMAN]->(mama:Person {user_id:$uid})\n            WITH r, copil, tata, mama,\n                 coalesce(c.adopted, false) AS adoptat, c.seq AS seq\n            WITH r, copil, adoptat, seq,\n                 [p IN [tata, mama] WHERE p IS NOT NULL] AS parinti\n            UNWIND parinti AS parinte\n            RETURN parinte.id AS source, copil.id AS target,\n                   r.id AS relation_id, seq AS seq,\n                   CASE WHEN adoptat THEN 'ADOPTIVE_PARENT' ELSE 'BIRTH_PARENT' END AS type\n            ",
            uid=uid,
        ))

        nodes = []
        for row in noduri:
            p = serializeaza_persoana(row["p"])
            p.pop("tel", None)
            p.pop("email_addr", None)
            nodes.append(p)

        edges = []
        perechi_vazute = set()
        for row in partner_rows:
            mid, fid = row["male_id"], row["female_id"]
            cheie = tuple(sorted([mid, fid]))
            if cheie in perechi_vazute:
                continue
            perechi_vazute.add(cheie)
            edges.append({
                "source": mid, "target": fid, "type": "PARTNER",
                "relation_id": row["relation_id"],
                "partner_type": row["partner_type"],
            })
        for row in parent_rows:
            edges.append(dict(row))

        return {"nodes": nodes, "edges": edges, "readonly": True}

class InviteRequest(BaseModel):
    email: str
    role: str = "editor"

class EmailMembruRequest(BaseModel):
    to_email: str
    to_name: str = ""
    subject: str
    message: str

class ParolaSchimbataRequest(BaseModel):
    email: str = ""
    full_name: str = ""

class RegisterRequest(BaseModel):
    email: str
    password: str
    full_name: str = ""

class ResendRequest(BaseModel):
    email: str

class VerifyRequest(BaseModel):
    token: str

class RequestResetRequest(BaseModel):
    email: str

class ResetPasswordRequest(BaseModel):
    token: str
    password: str

def _upsert_appuser(s, uid: str, email: str, full_name: str):
    s.run(
        "MERGE (u:AppUser {user_id:$uid}) SET u.email=$em, u.full_name=$nm",
        uid=uid, em=email or "", nm=full_name or email or "",
    )

@app.get("/api/trees")
def listeaza_arbori(cu0: dict = Depends(doar_autentificat)):
    caller = cu0["user_id"]
    email  = (cu0["email"] or "").lower()
    nume   = cu0["full_name"]

    with driver.session() as s:
        _upsert_appuser(s, caller, email, nume)

        n_prop = s.run(
            "MATCH (p:Person {user_id:$uid}) RETURN count(p) AS n", uid=caller
        ).single()["n"]
        arbori = [{
            "tree_id": caller, "role": "owner", "is_owner": True,
            "owner_name": nume, "persons": n_prop,
        }]

        membri = list(s.run(
            '\n            MATCH (m:Member)\n            WHERE m.user_id = $uid\n               OR (m.email IS NOT NULL AND toLower(m.email) = $em)\n            RETURN m.tree_id AS tid, m.role AS role, m.user_id AS muid, m.email AS memail\n            ',
            uid=caller, em=email,
        ))
        for r in membri:
            tid = r["tid"]
            if tid == caller:
                continue
            role = r["role"] or "editor"
            if role == "admin":
                role = "editor"
                s.run(
                    "MATCH (m:Member {tree_id:$tid}) WHERE m.user_id=$uid OR toLower(m.email)=$em "
                    "SET m.role='editor'",
                    tid=tid, uid=caller, em=email,
                )
            if not r["muid"] and r["memail"]:
                s.run(
                    "MATCH (m:Member {tree_id:$tid}) WHERE toLower(m.email)=$em "
                    "SET m.user_id=$uid, m.status='active'",
                    tid=tid, em=r["memail"].lower(), uid=caller,
                )
            owner = s.run(
                "MATCH (u:AppUser {user_id:$t}) RETURN u.full_name AS n", t=tid
            ).single()
            n_pers = s.run(
                "MATCH (p:Person {user_id:$t}) RETURN count(p) AS n", t=tid
            ).single()["n"]
            arbori.append({
                "tree_id": tid, "role": role, "is_owner": False,
                "owner_name": (owner["n"] if owner else "Arbore partajat"),
                "persons": n_pers,
            })
    return arbori

@app.get("/api/collab/link")
def info_link_colaborare(request: Request, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        r = s.run(
            "MATCH (cl:CollabLink {tree_id:$tid}) RETURN cl.token AS t",
            tid=cu["user_id"],
        ).single()
    if not r or not r["t"]:
        return {"token": None, "url": None}
    return {"token": r["t"], "url": f"{frontend_url(request)}/join/{r['t']}"}

@app.post("/api/collab/link")
def creeaza_link_colaborare(request: Request, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    token = secrets.token_urlsafe(16)
    with driver.session() as s:
        s.run(
            "MERGE (cl:CollabLink {tree_id:$tid}) SET cl.token=$t, cl.role='editor'",
            tid=cu["user_id"], t=token,
        )
        audit_log(s, cu, "create_collab_link", "collab_link", cu["user_id"])
    return {"token": token, "url": f"{frontend_url(request)}/join/{token}"}

@app.delete("/api/collab/link")
def revoca_link_colaborare(cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        s.run("MATCH (cl:CollabLink {tree_id:$tid}) DELETE cl", tid=cu["user_id"])
        audit_log(s, cu, "revoke_collab_link", "collab_link", cu["user_id"])
    return {"ok": True}

@app.post("/api/collab/join/{token}")
def alatura_te_arbore(token: str, cu0: dict = Depends(doar_autentificat)):
    caller = cu0["user_id"]
    email  = (cu0["email"] or "").lower()
    nume   = cu0["full_name"]

    with driver.session() as s:
        r = s.run(
            "MATCH (cl:CollabLink {token:$t}) RETURN cl.tree_id AS tid, cl.role AS role",
            t=token,
        ).single()
        if not r:
            raise HTTPException(404, "Link invalid sau dezactivat.")
        tid  = r["tid"]
        role = r["role"] if r["role"] in ("editor", "viewer") else "editor"

        if tid == caller:
            return {"tree_id": tid, "role": "owner", "message": "Acesta este chiar arborele tău."}

        _upsert_appuser(s, caller, email, nume)

        s.run(
            "\n            MERGE (m:Member {tree_id:$tid, email:$em})\n            SET m.user_id = $uid,\n                m.full_name = $nm,\n                m.status = 'active',\n                m.role = $role\n            ",
            tid=tid, em=email, uid=caller, nm=nume, role=role,
        )
    return {"tree_id": tid, "role": role}

@app.get("/api/members")
def listeaza_membri(cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        s.run(
            "MATCH (m:Member {tree_id:$tid, role:'admin'}) SET m.role='editor'",
            tid=cu["user_id"],
        )
        rows = s.run(
            '\n            MATCH (m:Member {tree_id:$tid})\n            RETURN m.email AS email, m.user_id AS user_id,\n                   m.full_name AS full_name, m.role AS role, m.status AS status\n            ORDER BY m.email\n            ',
            tid=cu["user_id"],
        ).data()
    return rows

@app.post("/api/members/invite")
def invita_membru(req: InviteRequest, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    tid   = cu["user_id"]
    email = (req.email or "").strip().lower()
    role  = req.role if req.role in ("editor", "viewer") else "editor"

    if not email or "@" not in email:
        raise HTTPException(400, "Adresă de email invalidă.")
    if email == (cu["email"] or "").lower():
        raise HTTPException(400, "Ești deja proprietarul acestui arbore.")

    with driver.session() as s:
        u = s.run(
            "MATCH (u:AppUser) WHERE toLower(u.email)=$em "
            "RETURN u.user_id AS uid, u.full_name AS nm",
            em=email,
        ).single()
        uid = u["uid"] if u else None
        nm  = u["nm"] if u else None
        s.run(
            "\n            MERGE (m:Member {tree_id:$tid, email:$em})\n            SET m.role = $role,\n                m.user_id = $uid,\n                m.full_name = $nm,\n                m.status = CASE WHEN $uid IS NULL THEN 'pending' ELSE 'active' END\n            ",
            tid=tid, em=email, role=role, uid=uid, nm=nm,
        )
        audit_log(s, cu, "invite_member", "member", email, dupa={"email": email, "role": role})

    invitator = cu.get("full_name") or cu.get("email") or "Cineva"
    rol_ro    = "editor (poate adăuga și modifica)" if role == "editor" else "vizualizare (doar citire)"
    trimis = trimite_email(
        to_email=email,
        subject=f"{invitator} te-a invitat în arborele genealogic",
        heading="Ai fost invitat în familie",
        icon="👨‍👩‍👧",
        message=(
            f"{invitator} te-a invitat să te alături arborelui genealogic al familiei, "
            f"cu rol de {rol_ro}.\n\n"
            f"Ca să ai acces, intră în aplicație și autentifică-te (sau creează-ți cont) "
            f"folosind exact această adresă de email: {email}.\n\n"
            f"Te așteptăm să descoperi și să completezi povestea familiei!"
        ),
        link=frontend_url(request),
        link_label="Deschide aplicația",
        from_name=invitator,
    )
    return {"email": email, "role": role,
            "status": "active" if uid else "pending", "email_trimis": trimis}

@app.delete("/api/members")
def sterge_membru(email: str, cu: dict = Depends(utilizator_curent)):
    cere_owner(cu)
    with driver.session() as s:
        s.run(
            "MATCH (m:Member {tree_id:$tid}) WHERE toLower(m.email)=toLower($em) DELETE m",
            tid=cu["user_id"], em=email or "",
        )
        audit_log(s, cu, "remove_member", "member", email or "")
    return {"ok": True}

@app.post("/api/email/send")
def trimite_email_membru(request: Request, req: EmailMembruRequest, cu: dict = Depends(utilizator_curent)):
    if cu["role"] not in ("owner", "editor"):
        raise HTTPException(403, "Nu ai dreptul să trimiți emailuri.")
    dest = (req.to_email or "").strip()
    if not dest or "@" not in dest:
        raise HTTPException(400, "Adresă de email invalidă.")
    if not req.subject.strip() or not req.message.strip():
        raise HTTPException(400, "Subiectul și mesajul sunt obligatorii.")

    expeditor = cu.get("full_name") or cu.get("email") or "Arbore Genealogic"
    trimis = trimite_email(
        to_email=dest,
        to_name=req.to_name.strip(),
        subject=req.subject.strip(),
        heading=f"Mesaj de la {expeditor}",
        icon="💌",
        message=req.message.strip(),
        from_name=expeditor,
    )
    if not trimis:
        raise HTTPException(502, "Emailul nu a putut fi trimis.")
    return {"ok": True}

@app.post("/api/email/password-changed")
def notifica_parola_schimbata(req: ParolaSchimbataRequest, cu0: dict = Depends(doar_autentificat)):
    dest = (cu0.get("email") or req.email or "").strip()
    if not dest or "@" not in dest:
        raise HTTPException(400, "Adresă de email invalidă.")
    nume = cu0.get("full_name") or req.full_name or dest
    cand = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    trimite_email(
        to_email=dest,
        to_name=nume,
        subject="Parola contului tău a fost schimbată",
        heading="Parola a fost schimbată",
        icon="🔒",
        message=(
            f"Te informăm că parola contului tău de Arbore Genealogic a fost "
            f"schimbată cu succes la {cand}.\n\n"
            f"Dacă tu ai făcut această modificare, nu trebuie să faci nimic.\n\n"
            f"Dacă NU recunoști această acțiune, resetează-ți imediat parola și "
            f"verifică securitatea adresei tale de email."
        ),
    )
    return {"ok": True}

@app.post("/api/auth/register")
def inregistrare(request: Request, req: RegisterRequest):
    email = (req.email or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Adresă de email invalidă.")
    if len(req.password or "") < 6:
        raise HTTPException(400, "Parola trebuie să aibă minim 6 caractere.")

    uid = supabase_admin_creeaza_user(email, req.password, req.full_name)
    with driver.session() as s:
        _upsert_appuser(s, uid, email, req.full_name)
    token = creeaza_token_email(uid, email, req.full_name, kind="verify")
    trimis = trimite_email_verificare(email, req.full_name, token, request=request)
    return {"ok": True, "email": email, "email_trimis": trimis}

@app.post("/api/auth/resend-verification")
def retrimite_verificare(request: Request, req: ResendRequest):
    email = (req.email or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Adresă de email invalidă.")
    with driver.session() as s:
        row = s.run(
            "MATCH (t:EmailToken {email:$em, kind:'verify'}) "
            "RETURN t.user_id AS uid, t.full_name AS nm ORDER BY t.created_at DESC LIMIT 1",
            em=email,
        ).single()
    if row and row["uid"]:
        token = creeaza_token_email(row["uid"], email, row["nm"] or email, kind="verify")
        trimite_email_verificare(email, row["nm"] or email, token, request=request)
    return {"ok": True}

@app.post("/api/auth/verify-email")
def verifica_email(request: Request, req: VerifyRequest):
    tok = (req.token or "").strip()
    if not tok:
        raise HTTPException(400, "Token lipsă.")
    acum = datetime.now(timezone.utc)
    with driver.session() as s:
        row = s.run(
            "MATCH (t:EmailToken {token:$tok, kind:'verify'}) "
            "RETURN t.user_id AS uid, t.email AS em, t.used AS used, t.expires_at AS exp",
            tok=tok,
        ).single()
        if not row:
            raise HTTPException(400, "Link de verificare invalid.")
        if row["used"]:
            return {"ok": True, "already": True}
        try:
            expirat = datetime.fromisoformat(row["exp"]) < acum
        except Exception:
            expirat = False
        if expirat:
            raise HTTPException(400, "Linkul de verificare a expirat. Cere unul nou.")

        if not supabase_admin_confirma_user(row["uid"]):
            raise HTTPException(502, "Confirmarea contului a eșuat. Încearcă din nou.")
        s.run("MATCH (t:EmailToken {token:$tok}) SET t.used=true", tok=tok)
    return {"ok": True, "email": row["em"]}

@app.post("/api/auth/request-password-reset")
def cere_resetare_parola(request: Request, req: RequestResetRequest):
    email = (req.email or "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(400, "Adresă de email invalidă.")
    uid = gaseste_user_id_dupa_email(email) or gaseste_user_id_supabase_dupa_email(email)
    if uid:
        token = creeaza_token_email(uid, email, "", kind="reset")
        link  = f"{frontend_url(request)}/reset-password?token={token}"
        trimite_email(
            to_email=email,
            subject="Resetare parolă — Arbore Genealogic",
            heading="Resetare parolă",
            icon="🔑",
            message=(
                "Am primit o cerere de resetare a parolei contului tău de Arbore Genealogic.\n\n"
                "Apasă pe butonul de mai jos ca să alegi o parolă nouă. "
                "Linkul este valabil 24 de ore.\n\n"
                "Dacă nu tu ai cerut resetarea, ignoră acest mesaj — parola rămâne neschimbată."
            ),
            link=link,
            link_label="Resetează parola",
        )
    return {"ok": True}

@app.post("/api/auth/reset-password")
def reseteaza_parola(request: Request, req: ResetPasswordRequest):
    tok = (req.token or "").strip()
    if not tok:
        raise HTTPException(400, "Token lipsă.")
    if len(req.password or "") < 6:
        raise HTTPException(400, "Parola trebuie să aibă minim 6 caractere.")
    acum = datetime.now(timezone.utc)
    with driver.session() as s:
        row = s.run(
            "MATCH (t:EmailToken {token:$tok, kind:'reset'}) "
            "RETURN t.user_id AS uid, t.email AS em, t.used AS used, t.expires_at AS exp",
            tok=tok,
        ).single()
        if not row:
            raise HTTPException(400, "Link de resetare invalid.")
        if row["used"]:
            raise HTTPException(400, "Acest link de resetare a fost deja folosit.")
        try:
            expirat = datetime.fromisoformat(row["exp"]) < acum
        except Exception:
            expirat = False
        if expirat:
            raise HTTPException(400, "Linkul de resetare a expirat. Cere unul nou.")

        if not supabase_admin_seteaza_parola(row["uid"], req.password):
            raise HTTPException(502, "Schimbarea parolei a eșuat. Încearcă din nou.")
        s.run("MATCH (t:EmailToken {token:$tok}) SET t.used=true", tok=tok)

    trimite_email(
        to_email=row["em"],
        subject="Parola contului tău a fost schimbată",
        heading="Parola a fost schimbată",
        icon="🔒",
        message=(
            "Parola contului tău de Arbore Genealogic a fost schimbată cu succes. "
            "Te poți autentifica acum cu noua parolă.\n\n"
            "Dacă nu tu ai făcut această modificare, contactează-ne imediat."
        ),
    )
    return {"ok": True, "email": row["em"]}

@app.get("/health")
def health():
    try:
        with driver.session() as s:
            np = s.run("MATCH (p:Person) RETURN count(p) AS n").single()["n"]
            nr = s.run("MATCH (r:Relation) RETURN count(r) AS n").single()["n"]
        return {"status": "ok", "persons": np, "relations": nr}
    except Exception as e:
        return {"status": "error", "detail": str(e)}

@app.get("/")
def root():
    return {"message": "Arbore Genealogic API"}
